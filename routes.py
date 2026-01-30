import os
import sys
import json
from flask import render_template, request, redirect, url_for, flash, session, jsonify, send_file, make_response
from app import app, db
from models import Classroom, Schedule, Incident, ScheduleRequest
from datetime import datetime, timedelta

# OpenAI integration
try:
    import openai
    OPENAI_AVAILABLE = True
    openai.api_key = os.environ.get("OPENAI_API_KEY")
except ImportError:
    OPENAI_AVAILABLE = False
    openai = None
try:
    import pytz
    PYTZ_AVAILABLE = True
except ImportError:
    PYTZ_AVAILABLE = False
    # Fallback timezone handling
    from datetime import timezone
    pytz = None
import io
from urllib.parse import urljoin
from werkzeug.utils import secure_filename
import uuid

# Import optional dependencies with error handling
try:
    from pdf_generator import generate_classroom_pdf, generate_general_report, generate_availability_report
    PDF_AVAILABLE = True
except ImportError as e:
    import logging
    logging.warning(f"PDF generation not available: {e}")
    generate_classroom_pdf = generate_general_report = generate_availability_report = None
    PDF_AVAILABLE = False

try:
    from qr_generator import generate_qr_code
except ImportError as e:
    import logging
    logging.warning(f"QR code generation not available: {e}")
    generate_qr_code = None

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    EXCEL_AVAILABLE = True
except ImportError as e:
    import logging
    logging.warning(f"Excel functionality not available: {e}")
    openpyxl = Font = Alignment = PatternFill = None
    EXCEL_AVAILABLE = False

ADMIN_PASSWORD = "senai103103"
# All files are now stored in PostgreSQL database, no local file storage
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}
ALLOWED_EXCEL_EXTENSIONS = {'xlsx', 'xls'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def allowed_excel_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXCEL_EXTENSIONS

def is_admin_authenticated():
    return session.get('admin_authenticated', False)

def require_admin_auth(f):
    def decorated_function(*args, **kwargs):
        if not is_admin_authenticated():
            flash('Acesso negado. Autenticação necessária.', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    decorated_function.__name__ = f.__name__
    return decorated_function

@app.route('/')
def index():
    classrooms = Classroom.query.all()
    return render_template('index.html', classrooms=classrooms)

@app.route('/classroom/<int:classroom_id>')
def classroom_detail(classroom_id):
    from datetime import datetime
    current_date = get_brazil_time().date()
    
    classroom = Classroom.query.get_or_404(classroom_id)
    
    # Only show active schedules where courses haven't ended yet
    schedules = Schedule.query.filter_by(classroom_id=classroom_id, is_active=True).filter(
        db.or_(
            Schedule.end_date == None,  # No end date specified
            Schedule.end_date >= current_date  # Course hasn't ended yet
        )
    ).all()
    
    # Get incidents for this classroom using raw SQL to avoid SQLAlchemy column issues
    incidents = []
    try:
        from sqlalchemy import text
        with db.engine.connect() as conn:
            # First ensure the column exists
            try:
                conn.execute(text("ALTER TABLE incident ADD COLUMN IF NOT EXISTS hidden_from_classroom BOOLEAN DEFAULT FALSE"))
                conn.commit()
            except:
                pass
            
            # Use safe SQL query
            result = conn.execute(text("""
                SELECT id, classroom_id, reporter_name, reporter_email, description, 
                       created_at, is_active, is_resolved, admin_response, response_date,
                       COALESCE(hidden_from_classroom, FALSE) as hidden_from_classroom
                FROM incident 
                WHERE classroom_id = :classroom_id 
                  AND is_active = TRUE 
                  AND COALESCE(hidden_from_classroom, FALSE) = FALSE
                ORDER BY created_at DESC
            """), {'classroom_id': classroom_id})
            
            incident_data = result.fetchall()
            
            # Convert to Incident-like objects for template compatibility
            class IncidentProxy:
                def __init__(self, row):
                    self.id = row[0]
                    self.classroom_id = row[1]
                    self.reporter_name = row[2]
                    self.reporter_email = row[3]
                    self.description = row[4]
                    self.created_at = row[5]
                    self.is_active = row[6]
                    self.is_resolved = row[7]
                    self.admin_response = row[8]
                    self.response_date = row[9]
                    self.hidden_from_classroom = row[10]
            
            incidents = [IncidentProxy(row) for row in incident_data]
            
    except Exception as e:
        import logging
        logging.error(f"Incident query error: {e}")
        incidents = []
    
    return render_template('classroom.html', classroom=classroom, schedules=schedules, incidents=incidents)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        password = request.form.get('password')
        if password == ADMIN_PASSWORD:
            session['admin_authenticated'] = True
            session.permanent = True
            app.permanent_session_lifetime = timedelta(hours=2)
            flash('Login realizado com sucesso!', 'success')
            next_page = request.args.get('next')
            return redirect(next_page or url_for('index'))
        else:
            flash('Senha incorreta!', 'error')
    
    return render_template('auth.html')

@app.route('/logout')
def logout():
    session.pop('admin_authenticated', None)
    flash('Logout realizado com sucesso!', 'success')
    return redirect(url_for('index'))

@app.route('/install')
def install_instructions():
    """Página com instruções para instalar o aplicativo em diferentes dispositivos"""
    return render_template('install_instructions.html')

@app.route('/static/sw.js')
def service_worker():
    """Serve the service worker with correct MIME type"""
    return send_file('static/sw.js', mimetype='application/javascript')

# Error handlers para prevenir crashes
@app.errorhandler(404)
def not_found_error(error):
    """Página personalizada para erro 404"""
    return render_template('errors/404.html'), 404

@app.errorhandler(500)
def internal_error(error):
    """Página personalizada para erro 500"""
    db.session.rollback()
    return render_template('errors/500.html'), 500

@app.errorhandler(403)
def forbidden_error(error):
    """Página personalizada para erro 403"""
    return render_template('errors/403.html'), 403

@app.route('/edit_classroom/<int:classroom_id>', methods=['GET', 'POST'])
@require_admin_auth
def edit_classroom(classroom_id):
    current_date = get_brazil_time().date()
    
    classroom = Classroom.query.get_or_404(classroom_id)
    
    # Only show active schedules where courses haven't ended yet
    schedules = Schedule.query.filter_by(classroom_id=classroom_id, is_active=True).filter(
        db.or_(
            Schedule.end_date == None,  # No end date specified
            Schedule.end_date >= current_date  # Course hasn't ended yet
        )
    ).all()
    
    import logging
    logging.debug(f"Edit classroom showing {len(schedules)} active/current schedules for classroom {classroom_id} (expired courses hidden)")
    
    if request.method == 'POST':
        try:
            classroom.name = request.form.get('name', '')
            classroom.capacity = int(request.form.get('capacity', 0))
            classroom.has_computers = 'has_computers' in request.form
            classroom.software = request.form.get('software', '')
            classroom.description = request.form.get('description', '')

            classroom.block = request.form.get('block', '')
            classroom.admin_password = request.form.get('admin_password', '')
            
            # Handle image upload with PostgreSQL storage
            if 'image' in request.files:
                file = request.files['image']
                if file and file.filename and file.filename != '' and allowed_file(file.filename):
                    filename = secure_filename(file.filename)
                    # Store file data in database
                    classroom.image_data = file.read()
                    classroom.image_mimetype = file.mimetype
                    classroom.image_filename = filename
            
            # Handle Excel file upload with PostgreSQL storage
            if 'excel_file' in request.files:
                excel_file = request.files['excel_file']
                if excel_file and excel_file.filename and excel_file.filename != '' and allowed_excel_file(excel_file.filename):
                    filename = secure_filename(excel_file.filename)
                    # Store file data in database
                    classroom.excel_data = excel_file.read()
                    classroom.excel_mimetype = excel_file.mimetype
                    classroom.excel_filename = filename
                    
            classroom.updated_at = datetime.utcnow()
            
            db.session.commit()
            flash('Sala atualizada com sucesso!', 'success')
            return redirect(url_for('classroom_detail', classroom_id=classroom_id))
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao atualizar sala: {str(e)}', 'error')
            return render_template('edit_classroom.html', classroom=classroom, schedules=schedules)
    
    return render_template('edit_classroom.html', classroom=classroom, schedules=schedules)


@app.route('/download_excel/<int:classroom_id>')
def download_excel(classroom_id):
    try:
        classroom = Classroom.query.get_or_404(classroom_id)
        
        if not classroom.excel_data:
            flash('Nenhum arquivo Excel disponível para esta sala.', 'error')
            return redirect(url_for('classroom_detail', classroom_id=classroom_id))
        
        safe_filename = f"{classroom.name.replace(' ', '_')}_patrimonio.xlsx"
        return send_file(
            io.BytesIO(classroom.excel_data),
            mimetype=classroom.excel_mimetype or 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=safe_filename
        )
    except Exception as e:
        flash(f'Erro ao baixar arquivo: {str(e)}', 'error')
        return redirect(url_for('classroom_detail', classroom_id=classroom_id))

@app.route('/image/<int:classroom_id>')
def serve_image(classroom_id):
    """Serve images from PostgreSQL database"""
    try:
        classroom = Classroom.query.get_or_404(classroom_id)
        
        if not classroom.image_data:
            # Return default image or 404
            from flask import abort
            abort(404)
        
        return send_file(
            io.BytesIO(classroom.image_data),
            mimetype=classroom.image_mimetype or 'image/jpeg'
        )
    except Exception as e:
        from flask import abort
        abort(404)

@app.route('/upload_excel/<int:classroom_id>', methods=['POST'])
@require_admin_auth
def upload_excel(classroom_id):
    classroom = Classroom.query.get_or_404(classroom_id)
    
    if 'excel_file' not in request.files:
        flash('Nenhum arquivo selecionado.', 'error')
        return redirect(url_for('edit_classroom', classroom_id=classroom_id))
    
    excel_file = request.files['excel_file']
    
    if excel_file.filename == '':
        flash('Nenhum arquivo selecionado.', 'error')
        return redirect(url_for('edit_classroom', classroom_id=classroom_id))
    
    if excel_file and excel_file.filename and allowed_excel_file(excel_file.filename):
        try:
            filename = secure_filename(excel_file.filename or '')
            
            # Store file data in database
            classroom.excel_data = excel_file.read()
            classroom.excel_mimetype = excel_file.mimetype
            classroom.excel_filename = filename
            classroom.updated_at = datetime.utcnow()
            db.session.commit()
            
            flash('Arquivo Excel carregado com sucesso!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao carregar arquivo: {str(e)}', 'error')
    else:
        flash('Formato de arquivo não permitido. Use apenas arquivos .xlsx ou .xls', 'error')
    
    return redirect(url_for('edit_classroom', classroom_id=classroom_id))

@app.route('/delete_schedule/<int:schedule_id>', methods=['POST'])
@require_admin_auth
def delete_schedule(schedule_id):
    schedule = Schedule.query.get_or_404(schedule_id)
    classroom_id = schedule.classroom_id
    
    try:
        db.session.delete(schedule)
        db.session.commit()
        flash('Horário removido com sucesso!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao remover horário: {str(e)}', 'error')
    
    return redirect(url_for('edit_classroom', classroom_id=classroom_id))

@app.route('/add_incident/<int:classroom_id>', methods=['POST'])
def add_incident(classroom_id):
    classroom = Classroom.query.get_or_404(classroom_id)
    
    try:
        reporter_name = request.form.get('reporter_name', '').strip()
        reporter_email = request.form.get('reporter_email', '').strip()
        description = request.form.get('description', '').strip()
        
        if not reporter_name or not reporter_email or not description:
            flash('Todos os campos são obrigatórios para registrar uma ocorrência.', 'error')
            return redirect(url_for('classroom_detail', classroom_id=classroom_id))
        
        # Simple incident creation using SQLAlchemy model
        brazil_time = get_brazil_time().replace(tzinfo=None)
        
        incident = Incident(
            classroom_id=classroom_id,
            reporter_name=reporter_name,
            reporter_email=reporter_email,
            description=description
        )
        incident.created_at = brazil_time
        db.session.add(incident)
        db.session.commit()
        flash('Ocorrência registrada com sucesso! A administração será notificada.', 'success')
        return redirect(url_for('classroom_detail', classroom_id=classroom_id))
        
    except Exception as e:
        db.session.rollback()
        import logging
        logging.error(f'Erro ao registrar ocorrência: {str(e)}')
        flash('Erro ao registrar ocorrência. Tente novamente.', 'error')
        return redirect(url_for('classroom_detail', classroom_id=classroom_id))

@app.route('/hide_incident_from_classroom/<int:incident_id>', methods=['POST'])
@require_admin_auth
def hide_incident_from_classroom(incident_id):
    incident = Incident.query.get_or_404(incident_id)
    classroom_id = incident.classroom_id
    
    try:
        # Check if hidden_from_classroom column exists
        from sqlalchemy import text
        column_exists = False
        try:
            with db.engine.connect() as conn:
                if 'postgresql' in str(db.engine.url) or 'postgres' in str(db.engine.url):
                    result = conn.execute(text("""
                        SELECT column_name FROM information_schema.columns 
                        WHERE table_name='incident' AND column_name='hidden_from_classroom'
                    """))
                    column_exists = result.fetchone() is not None
                else:
                    try:
                        conn.execute(text("SELECT hidden_from_classroom FROM incident LIMIT 1"))
                        column_exists = True
                    except:
                        column_exists = False
        except:
            column_exists = False
        
        # Hide incident using raw SQL with column creation if needed
        with db.engine.connect() as conn:
            # Ensure column exists first
            try:
                conn.execute(text("ALTER TABLE incident ADD COLUMN IF NOT EXISTS hidden_from_classroom BOOLEAN DEFAULT FALSE"))
                conn.commit()
            except:
                pass
                
            # Use raw SQL to avoid SQLAlchemy issues
            conn.execute(text("""
                UPDATE incident 
                SET hidden_from_classroom = true 
                WHERE id = :incident_id
            """), {'incident_id': incident_id})
            conn.commit()
            
        flash('Ocorrência removida da visualização da sala!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao ocultar ocorrência: {str(e)}', 'error')
    
    return redirect(url_for('classroom_detail', classroom_id=classroom_id))

@app.route('/delete_incident/<int:incident_id>', methods=['POST'])
@require_admin_auth
def delete_incident(incident_id):
    incident = Incident.query.get_or_404(incident_id)
    
    # Check where we're coming from to redirect properly
    referrer = request.form.get('referrer', 'incidents_management')
    
    try:
        db.session.delete(incident)
        db.session.commit()
        flash('Ocorrência excluída permanentemente!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir ocorrência: {str(e)}', 'error')
    
    if referrer == 'classroom':
        return redirect(url_for('classroom_detail', classroom_id=incident.classroom_id))
    else:
        return redirect(url_for('incidents_management'))

@app.route('/admin/migrate_db')
@require_admin_auth  
def migrate_database():
    """Rota para migrar banco de dados - adicionar colunas faltantes"""
    try:
        from sqlalchemy import text
        with db.engine.connect() as conn:
            # Adicionar hidden_from_classroom
            try:
                conn.execute(text("ALTER TABLE incident ADD COLUMN IF NOT EXISTS hidden_from_classroom BOOLEAN DEFAULT FALSE"))
                conn.commit()
                flash('✅ Migração concluída com sucesso! Coluna hidden_from_classroom adicionada.', 'success')
            except Exception as e:
                if 'already exists' in str(e) or 'duplicate' in str(e):
                    flash('✅ Coluna hidden_from_classroom já existe.', 'info')  
                else:
                    flash(f'❌ Erro na migração: {str(e)}', 'error')
                    
    except Exception as e:
        flash(f'❌ Erro na migração: {str(e)}', 'error')
        
    return redirect(url_for('incidents_management'))

@app.route('/incidents_management')
@require_admin_auth
def incidents_management():
    """Admin panel for managing all incidents with filters"""
    try:
        # Get filter parameters
        status_filter = request.args.get('status', '')
        reporter_filter = request.args.get('reporter', '')
        classroom_filter = request.args.get('classroom', '')
        
        # Use raw SQL to avoid SQLAlchemy model issues with missing columns
        from sqlalchemy import text
        
        # Use simplified SQL with COALESCE to handle missing column gracefully
        base_sql = """
            SELECT id, classroom_id, reporter_name, reporter_email, description, 
                   created_at, is_active, is_resolved, admin_response, response_date,
                   COALESCE(hidden_from_classroom, FALSE) as hidden_from_classroom
        """
        where_clause = "WHERE is_active = true AND COALESCE(hidden_from_classroom, FALSE) = FALSE"
        
        # Add filters to WHERE clause
        filter_conditions = []
        params = {}
        
        if status_filter == 'pending':
            filter_conditions.append("is_resolved = false")
        elif status_filter == 'resolved':
            filter_conditions.append("is_resolved = true")
        
        if reporter_filter:
            filter_conditions.append("LOWER(reporter_name) LIKE LOWER(:reporter_filter)")
            params['reporter_filter'] = f'%{reporter_filter}%'
        
        if classroom_filter:
            try:
                classroom_id = int(classroom_filter)
                filter_conditions.append("classroom_id = :classroom_id")
                params['classroom_id'] = classroom_id
            except ValueError:
                pass
        
        # Complete SQL query
        if filter_conditions:
            full_sql = f"{base_sql} FROM incident {where_clause} AND {' AND '.join(filter_conditions)} ORDER BY created_at DESC"
        else:
            full_sql = f"{base_sql} FROM incident {where_clause} ORDER BY created_at DESC"
        
        # Execute query and convert to incident objects
        with db.engine.connect() as conn:
            # Ensure column exists first
            try:
                conn.execute(text("ALTER TABLE incident ADD COLUMN IF NOT EXISTS hidden_from_classroom BOOLEAN DEFAULT FALSE"))
                conn.commit()
            except:
                pass
                
            result = conn.execute(text(full_sql), params)
            incident_rows = result.fetchall()
        
        # Convert to Incident-like objects for template compatibility
        class IncidentProxy:
            def __init__(self, row):
                self.id = row[0]
                self.classroom_id = row[1]
                self.reporter_name = row[2]
                self.reporter_email = row[3]
                self.description = row[4]
                self.created_at = row[5]
                self.is_active = row[6]
                self.is_resolved = row[7]
                self.admin_response = row[8]
                self.response_date = row[9]
                if len(row) > 10:
                    self.hidden_from_classroom = row[10]
                else:
                    self.hidden_from_classroom = False
                
                # Add classroom relationship
                self.classroom = Classroom.query.get(self.classroom_id)
        
        incidents = [IncidentProxy(row) for row in incident_rows]
        
        # Get counts using safe SQL with COALESCE
        with db.engine.connect() as conn:
            # Ensure column exists first
            try:
                conn.execute(text("ALTER TABLE incident ADD COLUMN IF NOT EXISTS hidden_from_classroom BOOLEAN DEFAULT FALSE"))
                conn.commit()
            except:
                pass
                
            pending_result = conn.execute(text("""
                SELECT COUNT(*) FROM incident 
                WHERE is_active = true AND is_resolved = false 
                AND COALESCE(hidden_from_classroom, FALSE) = FALSE
            """))
            resolved_result = conn.execute(text("""
                SELECT COUNT(*) FROM incident 
                WHERE is_active = true AND is_resolved = true 
                AND COALESCE(hidden_from_classroom, FALSE) = FALSE
            """))
            
            pending_count = pending_result.scalar()
            resolved_count = resolved_result.scalar()
        
        # Get classrooms and reporters using safe queries
        classrooms = Classroom.query.all()
        
        with db.engine.connect() as conn:
            reporters_result = conn.execute(text("""
                SELECT DISTINCT reporter_name FROM incident 
                WHERE is_active = true 
                AND COALESCE(hidden_from_classroom, FALSE) = FALSE
            """))
            reporters = [row[0] for row in reporters_result.fetchall()]
        
        return render_template('incidents_management.html', 
                             incidents=incidents, 
                             pending_count=pending_count, 
                             resolved_count=resolved_count,
                             classrooms=classrooms,
                             reporters=reporters,
                             current_filters={
                                 'status': status_filter,
                                 'reporter': reporter_filter,
                                 'classroom': classroom_filter
                             })
    except Exception as e:
        import logging
        logging.error(f'Erro na gestão de ocorrências: {str(e)}')
        flash(f'Erro ao carregar gestão de ocorrências: {str(e)}', 'error')
        return redirect(url_for('index'))

@app.route('/respond_incident/<int:incident_id>', methods=['POST'])
@require_admin_auth
def respond_incident(incident_id):
    """Admin response to an incident"""
    incident = Incident.query.get_or_404(incident_id)
    
    try:
        admin_response = request.form.get('admin_response', '').strip()
        mark_resolved = request.form.get('mark_resolved') == '1'
        
        if not admin_response:
            flash('A resposta não pode estar vazia.', 'error')
            return redirect(url_for('incidents_management'))
        
        incident.admin_response = admin_response
        incident.response_date = get_brazil_time().replace(tzinfo=None)
        
        if mark_resolved:
            incident.is_resolved = True
        
        db.session.commit()
        
        status_msg = "e marcada como resolvida" if mark_resolved else ""
        flash(f'Resposta enviada com sucesso {status_msg}!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao enviar resposta: {str(e)}', 'error')
    
    return redirect(url_for('incidents_management'))

@app.route('/resolve_incident/<int:incident_id>', methods=['POST'])
@require_admin_auth
def resolve_incident(incident_id):
    """Mark an incident as resolved"""
    incident = Incident.query.get_or_404(incident_id)
    
    try:
        incident.is_resolved = True
        incident.response_date = get_brazil_time().replace(tzinfo=None)
        db.session.commit()
        flash('Ocorrência marcada como resolvida!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao resolver ocorrência: {str(e)}', 'error')
    
    return redirect(url_for('incidents_management'))

@app.route('/incidents_pdf_report')
@require_admin_auth
def incidents_pdf_report():
    """Generate PDF report of incidents with filters"""
    try:
        # Get the same filters as incidents_management
        status_filter = request.args.get('status', '')
        reporter_filter = request.args.get('reporter', '')
        classroom_filter = request.args.get('classroom', '')
        
        # Use raw SQL for PDF report to avoid column issues
        from sqlalchemy import text
        
        # Use simplified SQL with COALESCE for PDF report
        where_clause = "WHERE is_active = true AND COALESCE(hidden_from_classroom, FALSE) = FALSE"
        
        # Add filters
        filter_conditions = []
        params = {}
        
        if status_filter == 'pending':
            filter_conditions.append("is_resolved = false")
        elif status_filter == 'resolved':
            filter_conditions.append("is_resolved = true")
        
        if reporter_filter:
            filter_conditions.append("LOWER(reporter_name) LIKE LOWER(:reporter_filter)")
            params['reporter_filter'] = f'%{reporter_filter}%'
        
        if classroom_filter:
            try:
                classroom_id = int(classroom_filter)
                filter_conditions.append("classroom_id = :classroom_id")
                params['classroom_id'] = classroom_id
            except ValueError:
                pass
        
        # Complete SQL
        if filter_conditions:
            full_sql = f"SELECT id FROM incident {where_clause} AND {' AND '.join(filter_conditions)} ORDER BY created_at DESC"
        else:
            full_sql = f"SELECT id FROM incident {where_clause} ORDER BY created_at DESC"
        
        # Get incident data directly to avoid SQLAlchemy issues
        with db.engine.connect() as conn:
            # Ensure column exists first
            try:
                conn.execute(text("ALTER TABLE incident ADD COLUMN IF NOT EXISTS hidden_from_classroom BOOLEAN DEFAULT FALSE"))
                conn.commit()
            except:
                pass
                
            # Get full incident data
            full_data_sql = f"""
                SELECT id, classroom_id, reporter_name, reporter_email, description, 
                       created_at, is_active, is_resolved, admin_response, response_date
                FROM incident {where_clause}
            """
            if filter_conditions:
                full_data_sql += f" AND {' AND '.join(filter_conditions)}"
            full_data_sql += " ORDER BY created_at DESC"
            
            result = conn.execute(text(full_data_sql), params)
            incident_data = result.fetchall()
        
        # Convert to incident-like objects
        class IncidentProxy:
            def __init__(self, row):
                self.id = row[0]
                self.classroom_id = row[1]
                self.reporter_name = row[2]
                self.reporter_email = row[3]
                self.description = row[4]
                self.created_at = row[5]
                self.is_active = row[6]
                self.is_resolved = row[7]
                self.admin_response = row[8]
                self.response_date = row[9]
                self.classroom = Classroom.query.get(self.classroom_id)
        
        incidents = [IncidentProxy(row) for row in incident_data]
        
        # Generate PDF using ReportLab
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        
        # Create BytesIO buffer
        buffer = io.BytesIO()
        
        # Create the PDF object
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
        
        # Container for the 'Flowable' objects
        elements = []
        
        # Define styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, spaceAfter=30, textColor=colors.HexColor('#1f2937'))
        subtitle_style = ParagraphStyle('CustomSubtitle', parent=styles['Heading2'], fontSize=12, spaceAfter=20, textColor=colors.HexColor('#374151'))
        normal_style = styles['Normal']
        
        # Add title
        title = Paragraph("Relatório de Ocorrências - SENAI Morvan Figueiredo", title_style)
        elements.append(title)
        
        # Add generation date
        generation_date = f"Gerado em: {get_brazil_time().strftime('%d/%m/%Y às %H:%M')}"
        date_para = Paragraph(generation_date, normal_style)
        elements.append(date_para)
        elements.append(Spacer(1, 12))
        
        # Add filter info if any
        filter_info = []
        if status_filter:
            filter_info.append(f"Status: {'Pendentes' if status_filter == 'pending' else 'Resolvidas'}")
        if reporter_filter:
            filter_info.append(f"Reportado por: {reporter_filter}")
        if classroom_filter:
            classroom = Classroom.query.get(int(classroom_filter))
            if classroom:
                filter_info.append(f"Sala: {classroom.name}")
        
        if filter_info:
            filter_text = "Filtros aplicados: " + ", ".join(filter_info)
            filter_para = Paragraph(filter_text, subtitle_style)
            elements.append(filter_para)
            elements.append(Spacer(1, 12))
        
        if incidents:
            # Create table data
            data = [['ID', 'Sala', 'Reportado por', 'Data', 'Status', 'Descrição']]
            
            for incident in incidents:
                status = 'Resolvida' if incident.is_resolved else 'Pendente'
                # Truncate description for table
                description = incident.description[:50] + '...' if len(incident.description) > 50 else incident.description
                data.append([
                    f"#{incident.id}",
                    incident.classroom.name,
                    incident.reporter_name,
                    incident.created_at.strftime('%d/%m/%Y') if incident.created_at else '',
                    status,
                    description
                ])
            
            # Create table
            table = Table(data, colWidths=[0.8*inch, 1.5*inch, 1.5*inch, 1*inch, 1*inch, 2.2*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f2937')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]))
            
            elements.append(table)
            elements.append(Spacer(1, 20))
            
            # Add detailed incidents
            detailed_title = Paragraph("Detalhes das Ocorrências", subtitle_style)
            elements.append(detailed_title)
            
            for incident in incidents:
                # Incident header
                incident_header = f"Ocorrência #{incident.id} - {incident.classroom.name}"
                header_para = Paragraph(incident_header, ParagraphStyle('IncidentHeader', parent=styles['Heading3'], fontSize=11, textColor=colors.HexColor('#1f2937')))
                elements.append(header_para)
                
                # Incident details
                details = f"""<b>Reportado por:</b> {incident.reporter_name} ({incident.reporter_email})<br/>
                <b>Data:</b> {incident.created_at.strftime('%d/%m/%Y às %H:%M') if incident.created_at else 'Não informada'}<br/>
                <b>Status:</b> {'Resolvida' if incident.is_resolved else 'Pendente'}<br/>
                <b>Descrição:</b> {incident.description}<br/>"""
                
                if incident.admin_response:
                    details += f"<b>Resposta do Admin:</b> {incident.admin_response}<br/>"
                    if incident.response_date:
                        details += f"<b>Data da Resposta:</b> {incident.response_date.strftime('%d/%m/%Y às %H:%M')}<br/>"
                
                details_para = Paragraph(details, normal_style)
                elements.append(details_para)
                elements.append(Spacer(1, 12))
        else:
            no_incidents = Paragraph("Nenhuma ocorrência encontrada com os filtros aplicados.", normal_style)
            elements.append(no_incidents)
        
        # Add summary
        total_incidents = len(incidents)
        pending_incidents = len([i for i in incidents if not i.is_resolved])
        resolved_incidents = len([i for i in incidents if i.is_resolved])
        
        summary = f"""<b>Resumo:</b><br/>
        Total de ocorrências: {total_incidents}<br/>
        Pendentes: {pending_incidents}<br/>
        Resolvidas: {resolved_incidents}"""
        
        summary_para = Paragraph(summary, subtitle_style)
        elements.append(Spacer(1, 20))
        elements.append(summary_para)
        
        # Build PDF
        doc.build(elements)
        
        # Get the value of the BytesIO buffer and create response
        pdf_data = buffer.getvalue()
        buffer.close()
        
        timestamp = get_brazil_time().strftime("%Y%m%d_%H%M%S")
        filename = f'relatorio_ocorrencias_{timestamp}.pdf'
        
        response = make_response(pdf_data)
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        
        return response
        
    except Exception as e:
        flash(f'Erro ao gerar relatório PDF: {str(e)}', 'error')
        return redirect(url_for('incidents_management'))


@app.route('/migrate_uploads_to_db')
@require_admin_auth
def migrate_uploads_to_db():
    """Migrate any remaining files from uploads folder to PostgreSQL database"""
    try:
        import os
        uploads_folder = 'static/uploads'
        migrated_count = 0
        
        if not os.path.exists(uploads_folder):
            flash('Pasta uploads não encontrada - todos os arquivos já estão no banco.', 'info')
            return redirect(url_for('dashboard'))
        
        # Get all classrooms that might have old file references
        classrooms = Classroom.query.all()
        
        for classroom in classrooms:
            # Check if classroom has image_filename but no image_data
            if classroom.image_filename and not classroom.image_data:
                old_image_path = os.path.join(uploads_folder, classroom.image_filename)
                if os.path.exists(old_image_path):
                    try:
                        with open(old_image_path, 'rb') as f:
                            classroom.image_data = f.read()
                            # Determine mimetype from extension
                            ext = classroom.image_filename.lower().split('.')[-1]
                            mime_map = {
                                'jpg': 'image/jpeg', 'jpeg': 'image/jpeg',
                                'png': 'image/png', 'gif': 'image/gif'
                            }
                            classroom.image_mimetype = mime_map.get(ext, 'image/jpeg')
                            migrated_count += 1
                    except Exception as e:
                        import logging
                        logging.error(f"Erro ao migrar imagem {classroom.image_filename}: {e}")
            
            # Check if classroom has excel_filename but no excel_data
            # Also check Excel files with any uploads pattern
            if classroom.excel_filename and not classroom.excel_data:
                old_excel_path = os.path.join(uploads_folder, classroom.excel_filename)
                if os.path.exists(old_excel_path):
                    try:
                        with open(old_excel_path, 'rb') as f:
                            classroom.excel_data = f.read()
                            classroom.excel_mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                            migrated_count += 1
                    except Exception as e:
                        import logging
                        logging.error(f"Erro ao migrar Excel {classroom.excel_filename}: {e}")
        
        db.session.commit()
        flash(f'Migração concluída! {migrated_count} arquivos movidos para o banco PostgreSQL.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erro durante migração: {str(e)}', 'error')
    
    return redirect(url_for('dashboard'))

@app.route('/add_classroom', methods=['GET', 'POST'])
@require_admin_auth
def add_classroom():
    if request.method == 'POST':
        try:
            # Handle image upload with PostgreSQL storage
            image_data = None
            image_mimetype = None
            image_filename = ''
            if 'image' in request.files:
                file = request.files['image']
                if file and file.filename and file.filename != '' and allowed_file(file.filename):
                    filename = secure_filename(file.filename)
                    image_data = file.read()
                    image_mimetype = file.mimetype
                    image_filename = filename
            
            classroom = Classroom(
                name=request.form.get('name', ''),
                capacity=int(request.form.get('capacity', 0)),
                has_computers='has_computers' in request.form,
                software=request.form.get('software', ''),
                description=request.form.get('description', ''),
                block=request.form.get('block', ''),
                image_filename=image_filename,
                admin_password=request.form.get('admin_password', '')
            )
            
            # Set image data after creation
            if image_data:
                classroom.image_data = image_data
                classroom.image_mimetype = image_mimetype
            
            db.session.add(classroom)
            db.session.commit()
            
            # Create initial schedules if provided
            initial_shift = request.form.get('initial_shift')
            if initial_shift and request.form.get('initial_course'):
                initial_days = request.form.getlist('initial_days')
                if initial_days:
                    # Process date fields
                    initial_start_date = None
                    initial_end_date = None
                    
                    start_date_value = request.form.get('initial_start_date')
                    if start_date_value and start_date_value.strip():
                        try:
                            from datetime import datetime
                            initial_start_date = datetime.strptime(start_date_value.strip(), '%Y-%m-%d').date()
                        except ValueError:
                            pass
                    
                    end_date_value = request.form.get('initial_end_date')
                    if end_date_value and end_date_value.strip():
                        try:
                            from datetime import datetime
                            initial_end_date = datetime.strptime(end_date_value.strip(), '%Y-%m-%d').date()
                        except ValueError:
                            pass
                    
                    for day in initial_days:
                        schedule = Schedule(
                            classroom_id=classroom.id,
                            day_of_week=int(day),
                            shift=initial_shift,
                            course_name=request.form.get('initial_course', ''),
                            instructor=request.form.get('initial_instructor', ''),
                            start_time=request.form.get('initial_start_time', ''),
                            end_time=request.form.get('initial_end_time', ''),
                            start_date=initial_start_date,
                            end_date=initial_end_date
                        )
                        db.session.add(schedule)
                    db.session.commit()
                    
                    # Enhanced success message with date info
                    date_info = ""
                    if initial_start_date and initial_end_date:
                        date_info = f" (período: {initial_start_date.strftime('%d/%m/%Y')} a {initial_end_date.strftime('%d/%m/%Y')})"
                    elif initial_start_date:
                        date_info = f" (início: {initial_start_date.strftime('%d/%m/%Y')})"
                    
                    flash(f'Sala adicionada com {len(initial_days)} horários iniciais{date_info}!', 'success')
                else:
                    flash('Sala adicionada com sucesso!', 'success')
            else:
                flash('Sala adicionada com sucesso!', 'success')
            
            return redirect(url_for('classroom_detail', classroom_id=classroom.id))
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao adicionar sala: {str(e)}', 'error')
            
    return render_template('edit_classroom.html', classroom=None)

@app.route('/schedule_management')
@require_admin_auth
def schedule_management():
    current_date = get_brazil_time().date()
    
    classrooms = Classroom.query.all()
    
    # Only show active schedules where courses haven't ended yet
    schedules = Schedule.query.filter_by(is_active=True).filter(
        db.or_(
            Schedule.end_date == None,  # No end date specified
            Schedule.end_date >= current_date  # Course hasn't ended yet
        )
    ).all()
    
    print(f"DEBUG: Schedule management showing {len(schedules)} active/current schedules (expired courses hidden)")
    
    # Organize schedules by classroom and day
    schedule_map = {}
    for schedule in schedules:
        if schedule.classroom_id not in schedule_map:
            schedule_map[schedule.classroom_id] = {}
        if schedule.day_of_week not in schedule_map[schedule.classroom_id]:
            schedule_map[schedule.classroom_id][schedule.day_of_week] = {}
        schedule_map[schedule.classroom_id][schedule.day_of_week][schedule.shift] = schedule
    
    return render_template('schedule_management.html', 
                         classrooms=classrooms, 
                         schedules=schedules,
                         schedule_map=schedule_map)

@app.route('/add_schedule', methods=['POST'])
@require_admin_auth
def add_schedule():
    try:
        classroom_id = int(request.form.get('classroom_id') or 0)
        days = request.form.getlist('days')
        
        # Handle single day submissions from first modal
        single_day = request.form.get('day_of_week')
        if single_day is not None and single_day != '':
            days = [single_day]
        
        shift = request.form.get('shift')
        course_name = request.form.get('course_name', '')
        instructor = request.form.get('instructor', '')
        start_time = request.form.get('start_time', '')
        end_time = request.form.get('end_time', '')
        start_date_str = request.form.get('start_date', '')
        end_date_str = request.form.get('end_date', '')
        
        # Parse dates with error handling
        start_date = None
        end_date = None
        if start_date_str and start_date_str.strip():
            try:
                start_date = datetime.strptime(start_date_str.strip(), '%Y-%m-%d').date()
            except ValueError:
                flash('Data de início inválida.', 'error')
                return redirect(url_for('schedule_management'))
        if end_date_str and end_date_str.strip():
            try:
                end_date = datetime.strptime(end_date_str.strip(), '%Y-%m-%d').date()
            except ValueError:
                flash('Data de fim inválida.', 'error')
                return redirect(url_for('schedule_management'))
        
        print(f"DEBUG: Adding schedule - classroom_id: {classroom_id}, days: {days}, shift: {shift}")
        
        created_count = 0
        existing_count = 0
        
        if not days or len(days) == 0:
            flash('Nenhum dia foi selecionado!', 'error')
            return redirect(url_for('schedule_management'))
        
        for day in days:
            day_int = int(day)
            
            # Check for overlaps taking dates into account
            existing_schedules = Schedule.query.filter_by(
                classroom_id=classroom_id,
                day_of_week=day_int,
                shift=shift,
                is_active=True
            ).all()
            
            has_overlap = False
            for existing in existing_schedules:
                # If both have dates, check for date range overlap
                if start_date and end_date and existing.start_date and existing.end_date:
                    # Overlap if: (StartA <= EndB) and (EndA >= StartB)
                    if start_date <= existing.end_date and end_date >= existing.start_date:
                        has_overlap = True
                        break
                # If one has no dates (permanent) and the other does, it's an overlap
                elif (not start_date or not end_date) or (not existing.start_date or not existing.end_date):
                    has_overlap = True
                    break
            
            print(f"DEBUG: Day {day_int}, has_overlap: {has_overlap}")
            
            if not has_overlap:
                schedule = Schedule(
                    classroom_id=classroom_id,
                    day_of_week=day_int,
                    shift=shift or '',
                    course_name=course_name,
                    instructor=instructor,
                    start_time=start_time,
                    end_time=end_time,
                    start_date=start_date,
                    end_date=end_date
                )
                db.session.add(schedule)
                created_count += 1
                print(f"DEBUG: Created schedule for day {day_int}")
            else:
                existing_count += 1
                print(f"DEBUG: Schedule already exists or overlaps for day {day_int}")
        
        if created_count > 0:
            db.session.commit()
            if existing_count > 0:
                flash(f'{created_count} horários adicionados, {existing_count} já existiam!', 'success')
            else:
                flash(f'{created_count} horários adicionados com sucesso!', 'success')
        elif existing_count > 0:
            flash(f'Todos os {existing_count} horários selecionados já existem!', 'warning')
        else:
            flash('Nenhum horário foi selecionado!', 'error')
        
    except Exception as e:
        print(f"DEBUG: Error in add_schedule: {str(e)}")
        db.session.rollback()
        flash(f'Erro ao adicionar horários: {str(e)}', 'error')
    
    return redirect(url_for('schedule_management'))



@app.route('/delete_classroom/<int:classroom_id>', methods=['POST'])
@require_admin_auth
def delete_classroom(classroom_id):
    try:
        classroom = Classroom.query.get_or_404(classroom_id)
        classroom_name = classroom.name
        
        # Delete all associated schedules and incidents
        Schedule.query.filter_by(classroom_id=classroom_id).delete()
        Incident.query.filter_by(classroom_id=classroom_id).delete()
        
        # Delete the classroom
        db.session.delete(classroom)
        db.session.commit()
        
        flash(f'Sala "{classroom_name}" excluída com sucesso!', 'success')
        return redirect(url_for('index'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir sala: {str(e)}', 'error')
        return redirect(url_for('edit_classroom', classroom_id=classroom_id))

@app.route('/dashboard')
def dashboard():
    # Get filter parameters
    block_filter = request.args.get('block', '')
    instructor_filter = request.args.get('instructor', '')
    software_filter = request.args.get('software', '')
    has_computers_filter = request.args.get('has_computers', '')
    capacity_filter = request.args.get('capacity', '')
    day_filter = request.args.get('day', '')
    shift_filter = request.args.get('shift', '')
    week_filter = request.args.get('week', '')  # New week filter parameter
    
    # Build classroom query with filters
    classroom_query = Classroom.query
    if block_filter:
        classroom_query = classroom_query.filter(Classroom.block.contains(block_filter))
    if software_filter:
        classroom_query = classroom_query.filter(Classroom.software.contains(software_filter))
    if has_computers_filter:
        has_computers_bool = has_computers_filter.lower() == 'true'
        classroom_query = classroom_query.filter(Classroom.has_computers == has_computers_bool)
    if capacity_filter:
        capacity_ranges = {
            'small': (0, 20),
            'medium': (21, 35),
            'large': (36, 100)
        }
        if capacity_filter in capacity_ranges:
            min_cap, max_cap = capacity_ranges[capacity_filter]
            classroom_query = classroom_query.filter(
                Classroom.capacity >= min_cap,
                Classroom.capacity <= max_cap
            )
    
    classrooms = classroom_query.all()
    
    # Build schedule query with filters - ONLY SHOW ACTIVE/CURRENT COURSES
    current_date = get_brazil_time().date()
    
    # Calculate week dates for filtering
    if week_filter:
        try:
            # Parse the week filter date (format: YYYY-MM-DD)
            week_start_date = datetime.strptime(week_filter, '%Y-%m-%d').date()
            # Get Monday of that week
            days_since_monday = week_start_date.weekday()
            week_monday = week_start_date - timedelta(days=days_since_monday)
            week_sunday = week_monday + timedelta(days=6)
        except (ValueError, TypeError):
            # If invalid date, use current week
            days_since_monday = current_date.weekday()
            week_monday = current_date - timedelta(days=days_since_monday)
            week_sunday = week_monday + timedelta(days=6)
    else:
        # Default to current week
        days_since_monday = current_date.weekday()
        week_monday = current_date - timedelta(days=days_since_monday)
        week_sunday = week_monday + timedelta(days=6)
    
    schedule_query = Schedule.query.filter_by(is_active=True)
    
    # Filter out expired courses - only show courses that haven't ended yet OR courses running in the selected week
    if week_filter:
        # For week filter, show courses that are active during the selected week
        schedule_query = schedule_query.filter(
            db.and_(
                db.or_(
                    Schedule.start_date.is_(None),  # No start date specified
                    Schedule.start_date <= week_sunday  # Course started before or during the week
                ),
                db.or_(
                    Schedule.end_date == None,  # No end date specified
                    Schedule.end_date >= week_monday  # Course ends after or during the week
                )
            )
        )
    else:
        # For normal view, only show courses that haven't ended yet
        schedule_query = schedule_query.filter(
            db.or_(
                Schedule.end_date == None,  # No end date specified
                Schedule.end_date >= current_date  # Course hasn't ended yet
            )
        )
    
    if day_filter:
        schedule_query = schedule_query.filter(Schedule.day_of_week == int(day_filter))
    if shift_filter:
        schedule_query = schedule_query.filter(Schedule.shift == shift_filter)
    if instructor_filter:
        schedule_query = schedule_query.filter(Schedule.instructor.ilike(f'%{instructor_filter}%'))
    
    schedules = schedule_query.all()
    print(f"DEBUG: Dashboard showing {len(schedules)} active/current schedules (expired courses hidden)")
    
    # Filter classrooms by instructor if specified
    if instructor_filter:
        classroom_ids_with_instructor = set(s.classroom_id for s in schedules)
        classroom_query = classroom_query.filter(Classroom.id.in_(classroom_ids_with_instructor))
    
    # Organize schedules by classroom and day
    schedule_map = {}
    for schedule in schedules:
        if schedule.classroom_id not in schedule_map:
            schedule_map[schedule.classroom_id] = {}
        if schedule.day_of_week not in schedule_map[schedule.classroom_id]:
            schedule_map[schedule.classroom_id][schedule.day_of_week] = {}
        schedule_map[schedule.classroom_id][schedule.day_of_week][schedule.shift] = schedule
    
    # Calculate statistics
    total_slots = len(classrooms) * 23  # 6 days * 4 shifts - 1 (no Saturday night)
    occupied_slots = len([s for s in schedules if s.classroom_id in [c.id for c in classrooms]])
    free_slots = total_slots - occupied_slots
    occupancy_rate = (occupied_slots / total_slots * 100) if total_slots > 0 else 0
    
    # Get unique filter options
    all_classrooms = Classroom.query.all()
    blocks = sorted(list(set(c.block for c in all_classrooms if c.block)))
    all_schedules = Schedule.query.filter_by(is_active=True).all()
    instructors = sorted(list(set(s.instructor for s in all_schedules if s.instructor and s.instructor.strip())))
    software_list = sorted(list(set(software.strip() for c in all_classrooms if c.software for software in c.software.split(',') if software.strip())))
    
    return render_template('dashboard.html', 
                         classrooms=classrooms, 
                         schedule_map=schedule_map,
                         free_slots=free_slots,
                         occupied_slots=occupied_slots,
                         occupancy_rate=occupancy_rate,
                         blocks=blocks,
                         instructors=instructors,
                         software_list=software_list,
                         current_filters={
                             'block': block_filter,
                             'instructor': instructor_filter,
                             'software': software_filter,
                             'has_computers': has_computers_filter,
                             'capacity': capacity_filter,
                             'day': day_filter,
                             'shift': shift_filter,
                             'week': week_filter
                         },
                         week_dates={
                             'monday': week_monday,
                             'sunday': week_sunday,
                             'formatted': f"{week_monday.strftime('%d/%m')} - {week_sunday.strftime('%d/%m/%Y')}"
                         })

@app.route('/availability')
def availability():
    return redirect(url_for('dashboard'))

def get_brazil_time():
    """Get current time in Brazil timezone (UTC-3)"""
    try:
        if PYTZ_AVAILABLE and pytz and 'pytz' in globals():
            brazil_tz = pytz.timezone('America/Sao_Paulo')
            return datetime.now(brazil_tz)
        else:
            # Fallback: subtract 3 hours from UTC to approximate Brazil time
            utc_time = datetime.utcnow()
            return utc_time - timedelta(hours=3)
    except Exception:
        # Ultimate fallback - just use UTC
        return datetime.utcnow()

def get_current_shift():
    """Get the current shift based on Brazil time"""
    now = get_brazil_time()
    current_hour = now.hour
    current_minute = now.minute
    current_time_minutes = current_hour * 60 + current_minute
    
    # Use real current time
    # current_hour = 14  # FOR TESTING only
    # current_minute = 30
    # current_time_minutes = current_hour * 60 + current_minute
    
    print(f"DEBUG: Current time: {current_hour}:{current_minute:02d} ({current_time_minutes} minutes)")
    
    # Define shift time ranges in minutes
    # Morning: 7:30-12:00 (450-720 minutes)
    # Afternoon: 13:00-18:00 (780-1080 minutes)  
    # Night: 18:30-22:30 (1110-1350 minutes)
    # Fullday: 8:00-17:00 (480-1020 minutes) - as per schedule data
    
    current_shifts = []
    
    if 450 <= current_time_minutes <= 720:  # Morning
        current_shifts.append('morning')
        print(f"DEBUG: Added morning shift")
    if 780 <= current_time_minutes <= 1080:  # Afternoon
        current_shifts.append('afternoon')
        print(f"DEBUG: Added afternoon shift")
    if 1110 <= current_time_minutes <= 1350:  # Night
        current_shifts.append('night')
        print(f"DEBUG: Added night shift")
    # Fullday: check if within fullday hours (8:00-17:00)
    if 480 <= current_time_minutes <= 1020:  # Fullday
        current_shifts.append('fullday')
        print(f"DEBUG: Added fullday shift")
        
    print(f"DEBUG: Current active shifts: {current_shifts}")
    return current_shifts

def get_availability_for_date(target_date=None, shift_filter=None):
    """Helper function to get room availability for a specific date and optional shift"""
    if target_date is None:
        target_date = get_brazil_time()
    
    # Get day of week (0=Monday, 6=Sunday)
    target_day = target_date.weekday()
    target_date_only = target_date.date()
    
    classrooms = Classroom.query.all()
    
    print(f"DEBUG: Checking availability for date: {target_date_only}, day of week: {target_day}")
    
    # Check if it's Sunday
    if target_day == 6:  # Sunday
        return {
            'available_rooms': classrooms,
            'occupied_rooms': [],
            'period_description': "Domingo - Escola fechada",
            'total_rooms': len(classrooms)
        }
    
    # If no shift filter is provided and we're checking current time, get current shifts
    if shift_filter is None or shift_filter == 'all':
        # If checking current date, determine the primary current shift
        if target_date.date() == get_brazil_time().date():
            current_shifts = get_current_shift()
            print(f"DEBUG: Checking current date, active shifts: {current_shifts}")
            
            if not current_shifts:  # Outside operating hours
                return {
                    'available_rooms': classrooms,
                    'occupied_rooms': [],
                    'period_description': "Fora do horário de funcionamento",
                    'total_rooms': len(classrooms)
                }
            
            # Determine which shift to check based on current time
            # Priority: specific shifts (morning, afternoon, night) over fullday
            primary_shift = None
            
            # Check for specific time-based shifts first
            if 'morning' in current_shifts:
                primary_shift = 'morning'
            elif 'afternoon' in current_shifts:
                primary_shift = 'afternoon'
            elif 'night' in current_shifts:
                primary_shift = 'night'
            elif 'fullday' in current_shifts:
                primary_shift = 'fullday'
            
            occupied_schedules = []
            
            if primary_shift:
                # Get schedules for the primary shift only - PRECISE DATE CHECKING
                all_schedules = Schedule.query.filter_by(
                    day_of_week=target_day,
                    shift=primary_shift,
                    is_active=True
                ).all()
                
                # Filter by actual course dates
                active_schedules = []
                for schedule in all_schedules:
                    if schedule.start_date and schedule.end_date:
                        if schedule.start_date <= target_date_only <= schedule.end_date:
                            active_schedules.append(schedule)
                            print(f"DEBUG: Schedule {schedule.id} ({schedule.shift}) is ACTIVE (course runs {schedule.start_date} to {schedule.end_date})")
                        else:
                            print(f"DEBUG: Schedule {schedule.id} ({schedule.shift}) is EXPIRED/FUTURE (course runs {schedule.start_date} to {schedule.end_date}, today is {target_date_only})")
                    else:
                        # If no dates specified, consider it active (backward compatibility)
                        active_schedules.append(schedule)
                        print(f"DEBUG: Schedule {schedule.id} ({schedule.shift}) has no date restrictions, treating as active")
                
                occupied_schedules.extend(active_schedules)
                print(f"DEBUG: Using primary shift '{primary_shift}', found {len(active_schedules)} ACTIVE schedules out of {len(all_schedules)} total")
                
                # CRITICAL LOGIC: Only add fullday schedules if we're checking for CURRENT time
                # This prevents fullday classes from appearing when user filters by specific shift
                if primary_shift in ['morning', 'afternoon'] and target_date.date() == get_brazil_time().date():
                    all_fullday_schedules = Schedule.query.filter_by(
                        day_of_week=target_day,
                        shift='fullday',
                        is_active=True
                    ).all()
                    
                    active_fullday_schedules = []
                    for schedule in all_fullday_schedules:
                        if schedule.start_date and schedule.end_date:
                            if schedule.start_date <= target_date_only <= schedule.end_date:
                                active_fullday_schedules.append(schedule)
                                print(f"DEBUG: Fullday schedule {schedule.id} is ACTIVE (overlaps with current {primary_shift} shift)")
                            else:
                                print(f"DEBUG: Fullday schedule {schedule.id} is EXPIRED/FUTURE")
                        else:
                            active_fullday_schedules.append(schedule)
                    
                    occupied_schedules.extend(active_fullday_schedules)
                    print(f"DEBUG: Added {len(active_fullday_schedules)} ACTIVE fullday schedules for CURRENT TIME overlap")
                else:
                    print(f"DEBUG: Skipping fullday overlap - not checking current time or not morning/afternoon shift")
        else:
            # For other dates (future/past), check ALL shifts to get complete availability picture
            print(f"DEBUG: Checking NON-CURRENT date {target_date_only} - checking ALL shifts for complete availability")
            all_schedules = Schedule.query.filter_by(day_of_week=target_day, is_active=True).all()
            
            active_schedules = []
            for schedule in all_schedules:
                if schedule.start_date and schedule.end_date:
                    if schedule.start_date <= target_date_only <= schedule.end_date:
                        active_schedules.append(schedule)
                        print(f"DEBUG: Schedule {schedule.id} ({schedule.shift} - {schedule.course_name}) is ACTIVE on {target_date_only}")
                    else:
                        print(f"DEBUG: Schedule {schedule.id} ({schedule.shift} - {schedule.course_name}) is EXPIRED/FUTURE on {target_date_only}")
                else:
                    active_schedules.append(schedule)
                    print(f"DEBUG: Schedule {schedule.id} ({schedule.shift} - {schedule.course_name}) has no date restrictions, treating as active")
            
            occupied_schedules = active_schedules
            print(f"DEBUG: Future/past date check - found {len(active_schedules)} ACTIVE schedules out of {len(all_schedules)} total")
    else:
        # Apply specific shift filter - ULTRA PRECISE: ONLY show rooms occupied by EXACTLY that shift
        print(f"DEBUG: PRECISE FILTER MODE - Looking for shift '{shift_filter}' on {target_date_only}")
        
        # Get schedules that EXACTLY match the requested shift
        all_schedules = Schedule.query.filter_by(
            day_of_week=target_day,
            shift=shift_filter,
            is_active=True
        ).all()
        
        active_schedules = []
        for schedule in all_schedules:
            if schedule.start_date and schedule.end_date:
                if schedule.start_date <= target_date_only <= schedule.end_date:
                    active_schedules.append(schedule)
                    print(f"DEBUG: Schedule {schedule.id} ({schedule.shift} - {schedule.course_name}) is ACTIVE with exact shift filter '{shift_filter}'")
                else:
                    print(f"DEBUG: Schedule {schedule.id} ({schedule.shift} - {schedule.course_name}) is EXPIRED/FUTURE with shift filter")
            else:
                active_schedules.append(schedule)
                print(f"DEBUG: Schedule {schedule.id} ({schedule.shift} - {schedule.course_name}) has no date restrictions, treating as active for shift filter")
        
        occupied_schedules = active_schedules
        
        # CRITICAL: For specific shift filters, we need to consider fullday classes as conflicts too
        # BUT only when the user is NOT specifically looking for fullday
        if shift_filter != 'fullday':
            print(f"DEBUG: Checking if any fullday classes conflict with '{shift_filter}' filter")
            
            # Get fullday schedules for this day
            all_fullday_schedules = Schedule.query.filter_by(
                day_of_week=target_day,
                shift='fullday',
                is_active=True
            ).all()
            
            active_fullday_schedules = []
            for schedule in all_fullday_schedules:
                if schedule.start_date and schedule.end_date:
                    if schedule.start_date <= target_date_only <= schedule.end_date:
                        active_fullday_schedules.append(schedule)
                        print(f"DEBUG: Fullday schedule {schedule.id} ({schedule.course_name}) CONFLICTS with '{shift_filter}' filter")
                    else:
                        print(f"DEBUG: Fullday schedule {schedule.id} ({schedule.course_name}) is EXPIRED/FUTURE, no conflict")
                else:
                    active_fullday_schedules.append(schedule)
                    print(f"DEBUG: Fullday schedule {schedule.id} ({schedule.course_name}) has no date restrictions, treating as conflict")
            
            occupied_schedules.extend(active_fullday_schedules)
            print(f"DEBUG: Added {len(active_fullday_schedules)} conflicting fullday schedules to '{shift_filter}' filter")
        
        print(f"DEBUG: EXACT SHIFT FILTER '{shift_filter}' - found {len(active_schedules)} exact matches + {len(occupied_schedules) - len(active_schedules)} conflicting schedules")
    
    occupied_classroom_ids = set(schedule.classroom_id for schedule in occupied_schedules)
    
    available_rooms = [room for room in classrooms if room.id not in occupied_classroom_ids]
    occupied_rooms = [room for room in classrooms if room.id in occupied_classroom_ids]
    
    # Build period description
    days = ['Segunda', 'Terça', 'Quarta', 'Quinta', 'Sexta', 'Sábado', 'Domingo']
    day_name = days[target_day]
    
    if shift_filter and shift_filter != 'all':
        shift_names = {'morning': 'Manhã', 'afternoon': 'Tarde', 'fullday': 'Integral', 'night': 'Noite'}
        period_description = f"{day_name} - {shift_names.get(shift_filter, shift_filter)} (Filtro Específico)"
    elif target_date.date() == get_brazil_time().date() and (shift_filter is None or shift_filter == 'all'):
        # Show current period
        current_shifts = get_current_shift()
        if current_shifts:
            shift_names = {'morning': 'Manhã', 'afternoon': 'Tarde', 'fullday': 'Integral', 'night': 'Noite'}
            active_shift_names = [shift_names.get(shift, shift) for shift in current_shifts]
            period_description = f"{day_name} - {', '.join(set(active_shift_names))} (Agora)"
        else:
            period_description = f"{day_name} - Fora do horário"
    else:
        period_description = f"{day_name} - Todos os turnos"
    
    return {
        'available_rooms': available_rooms,
        'occupied_rooms': occupied_rooms,
        'occupied_schedules': occupied_schedules,
        'period_description': period_description,
        'total_rooms': len(classrooms)
    }

@app.route('/available_now')
def available_now():
    # Get query parameters for date and shift filtering
    date_param = request.args.get('date')
    shift_param = request.args.get('shift', 'all')
    
    # Parse target date
    if date_param:
        try:
            from datetime import datetime
            target_date = datetime.strptime(date_param, '%Y-%m-%d')
        except ValueError:
            target_date = get_brazil_time()
    else:
        target_date = get_brazil_time()
    
    # Get availability data
    availability_data = get_availability_for_date(target_date, shift_param)
    
    # Format date for display
    formatted_date = target_date.strftime('%d/%m/%Y')
    
    return render_template('available_now.html', 
                         available_rooms=availability_data['available_rooms'],
                         occupied_rooms=availability_data.get('occupied_rooms', []),
                         occupied_schedules=availability_data.get('occupied_schedules', []),
                         current_period=availability_data['period_description'],
                         total_rooms=availability_data['total_rooms'],
                         selected_date=formatted_date,
                         selected_date_iso=target_date.strftime('%Y-%m-%d'),
                         selected_shift=shift_param)

@app.route('/generate_pdf/<int:classroom_id>')
def generate_pdf(classroom_id):
    if not generate_classroom_pdf:
        flash('Geração de PDF não está disponível no momento.', 'error')
        return redirect(url_for('classroom_detail', classroom_id=classroom_id))
        
    classroom = Classroom.query.get_or_404(classroom_id)
    schedules = Schedule.query.filter_by(classroom_id=classroom_id, is_active=True).all()
    
    try:
        pdf_buffer = generate_classroom_pdf(classroom, schedules)
        
        return send_file(
            io.BytesIO(pdf_buffer.getvalue()),
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'sala_{classroom.name.replace(" ", "_")}.pdf'
        )
    except Exception as e:
        flash(f'Erro ao gerar PDF: {str(e)}', 'error')
        return redirect(url_for('classroom_detail', classroom_id=classroom_id))

@app.route('/generate_general_report')
def generate_general_report_route():
    if not generate_general_report:
        flash('Geração de relatórios não está disponível no momento.', 'error')
        return redirect(url_for('dashboard'))
        
    try:
        classrooms = Classroom.query.all()
        schedules = Schedule.query.filter_by(is_active=True).all()
        
        pdf_buffer = generate_general_report(classrooms, schedules)
        
        return send_file(
            io.BytesIO(pdf_buffer.getvalue()),
            mimetype='application/pdf',
            as_attachment=True,
            download_name='relatorio_geral.pdf'
        )
    except Exception as e:
        flash(f'Erro ao gerar relatório: {str(e)}', 'error')
        return redirect(url_for('dashboard'))

@app.route('/generate_availability_report')
def generate_availability_report_route():
    if not generate_availability_report:
        flash('Geração de relatórios não está disponível no momento.', 'error')
        return redirect(url_for('dashboard'))
        
    try:
        classrooms = Classroom.query.all()
        schedules = Schedule.query.filter_by(is_active=True).all()
        
        pdf_buffer = generate_availability_report(classrooms, schedules)
        
        return send_file(
            io.BytesIO(pdf_buffer.getvalue()),
            mimetype='application/pdf',
            as_attachment=True,
            download_name='relatorio_disponibilidade.pdf'
        )
    except Exception as e:
        flash(f'Erro ao gerar relatório: {str(e)}', 'error')
        return redirect(url_for('dashboard'))

@app.route('/generate_qr/<int:classroom_id>')
def generate_qr(classroom_id):
    try:
        classroom = Classroom.query.get_or_404(classroom_id)
        
        if not generate_qr_code:
            flash('Geração de QR code não está disponível no momento.', 'error')
            return redirect(url_for('classroom_detail', classroom_id=classroom_id))
        
        # Generate the full URL for the classroom
        classroom_url = request.url_root.rstrip('/') + url_for('classroom_detail', classroom_id=classroom_id)
        
        qr_buffer = generate_qr_code(classroom_url, classroom.name)
        safe_filename = f'qr_sala_{classroom.name.replace(" ", "_").replace("/", "_")}.png'
        
        return send_file(
            io.BytesIO(qr_buffer.getvalue()),
            mimetype='image/png',
            as_attachment=True,
            download_name=safe_filename
        )
    except Exception as e:
        flash(f'Erro ao gerar QR code: {str(e)}', 'error')
        return redirect(url_for('classroom_detail', classroom_id=classroom_id))

# Exportação para Excel - versão corrigida
@app.route('/export_excel')
def export_excel():
    try:
        # Check if openpyxl is available
        if not openpyxl:
            flash('Funcionalidade de Excel não está disponível no momento.', 'error')
            return redirect(url_for('dashboard'))
            
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        
        # Sheet 1: Classrooms
        ws1 = wb.active
        ws1.title = "Salas de Aula"
        
        # Import openpyxl styles if available
        try:
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            Font = PatternFill = Alignment = None
        
        # Headers for classrooms
        headers1 = ['ID', 'Nome', 'Capacidade', 'Bloco', 'Tem Computadores', 'Softwares', 'Descrição']
        for col, header in enumerate(headers1, 1):
            cell = ws1.cell(row=1, column=col)
            if cell:
                cell.value = header
                if Font:
                    cell.font = Font(bold=True, color='FFFFFF')
                if PatternFill:
                    cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                if Alignment:
                    cell.alignment = Alignment(horizontal='center')
        
        # Data for classrooms
        classrooms = Classroom.query.all()
        for row, classroom in enumerate(classrooms, 2):
            ws1.cell(row=row, column=1).value = classroom.id
            ws1.cell(row=row, column=2).value = classroom.name
            ws1.cell(row=row, column=3).value = classroom.capacity
            ws1.cell(row=row, column=4).value = classroom.block
            ws1.cell(row=row, column=5).value = 'Sim' if classroom.has_computers else 'Não'
            ws1.cell(row=row, column=6).value = classroom.software
            ws1.cell(row=row, column=7).value = classroom.description
        
        # Auto-fit columns with safer approach
        try:
            for column_cells in ws1.columns:
                if column_cells and len(column_cells) > 0:
                    max_length = 0
                    column_letter = column_cells[0].column_letter
                    for cell in column_cells:
                        try:
                            if cell and cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    if ws1.column_dimensions and column_letter:
                        ws1.column_dimensions[column_letter].width = adjusted_width
        except Exception:
            pass  # Skip column sizing if there are issues
        
        # Sheet 2: Schedules
        ws2 = wb.create_sheet(title="Horários")
        
        # Headers for schedules
        headers2 = ['ID', 'Sala', 'Dia da Semana', 'Turno', 'Curso', 'Professor', 'Início', 'Fim', 'Ativo']
        for col, header in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=col)
            if cell:
                cell.value = header
                if Font:
                    cell.font = Font(bold=True, color='FFFFFF')
                if PatternFill:
                    cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                if Alignment:
                    cell.alignment = Alignment(horizontal='center')
        
        # Data for schedules
        schedules = Schedule.query.all()
        days = ['Segunda', 'Terça', 'Quarta', 'Quinta', 'Sexta', 'Sábado', 'Domingo']
        shifts = {'morning': 'Manhã', 'afternoon': 'Tarde', 'fullday': 'Integral', 'night': 'Noite'}
        
        for row, schedule in enumerate(schedules, 2):
            classroom = Classroom.query.get(schedule.classroom_id)
            ws2.cell(row=row, column=1).value = schedule.id
            ws2.cell(row=row, column=2).value = classroom.name if classroom else 'N/A'
            ws2.cell(row=row, column=3).value = days[schedule.day_of_week]
            ws2.cell(row=row, column=4).value = shifts.get(schedule.shift, schedule.shift)
            ws2.cell(row=row, column=5).value = schedule.course_name
            ws2.cell(row=row, column=6).value = schedule.instructor
            ws2.cell(row=row, column=7).value = schedule.start_time
            ws2.cell(row=row, column=8).value = schedule.end_time
            ws2.cell(row=row, column=9).value = 'Sim' if schedule.is_active else 'Não'
        
        # Auto-fit columns for sheet 2
        for column_cells in ws2.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws2.column_dimensions[column_letter].width = adjusted_width
        
        # Sheet 3: Statistics
        ws3 = wb.create_sheet(title="Estatísticas")
        ws3.cell(row=1, column=1).value = "Estatística"
        ws3.cell(row=1, column=2).value = "Valor"
        
        # Style headers
        for col in [1, 2]:
            cell = ws3.cell(row=1, column=col)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Statistics data
        total_classrooms = len(classrooms)
        total_schedules = len([s for s in schedules if s.is_active])
        total_slots = total_classrooms * 23  # 6 days * 4 shifts - 1 (no Saturday night)
        occupancy_rate = (total_schedules / total_slots * 100) if total_slots > 0 else 0
        
        stats_data = [
            ['Total de Salas', total_classrooms],
            ['Total de Horários Ativos', total_schedules],
            ['Taxa de Ocupação (%)', f"{occupancy_rate:.1f}%"],
            ['Salas com Computadores', len([c for c in classrooms if c.has_computers])],
            ['Salas sem Computadores', len([c for c in classrooms if not c.has_computers])]
        ]
        
        for row, (stat, value) in enumerate(stats_data, 2):
            ws3.cell(row=row, column=1).value = stat
            ws3.cell(row=row, column=2).value = value
        
        # Auto-fit columns for sheet 3
        for column_cells in ws3.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws3.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f'relatorio_senai_{timestamp}.xlsx'
        
        response = make_response(output.getvalue())
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        
        return response
    except Exception as e:
        flash(f'Erro ao gerar Excel: {str(e)}', 'error')
        return redirect(url_for('dashboard'))

@app.route('/export_filtered_excel')
def export_filtered_excel():
    try:
        # Get the same filters as dashboard
        block_filter = request.args.get('block', '')
        # Remove floor filter as it doesn't exist in the model
        has_computers_filter = request.args.get('has_computers', '')
        capacity_filter = request.args.get('capacity', '')
        day_filter = request.args.get('day', '')
        shift_filter = request.args.get('shift', '')
        
        # Build filtered queries
        classroom_query = Classroom.query
        if block_filter:
            classroom_query = classroom_query.filter(Classroom.block == block_filter)
        # Remove floor filter as it doesn't exist in the model
        if has_computers_filter:
            has_computers_bool = has_computers_filter.lower() == 'true'
            classroom_query = classroom_query.filter(Classroom.has_computers == has_computers_bool)
        if capacity_filter:
            capacity_ranges = {
                'small': (0, 20),
                'medium': (21, 35),
                'large': (36, 100)
            }
            if capacity_filter in capacity_ranges:
                min_cap, max_cap = capacity_ranges[capacity_filter]
                classroom_query = classroom_query.filter(
                    Classroom.capacity >= min_cap,
                    Classroom.capacity <= max_cap
                )
        
        filtered_classrooms = classroom_query.all()
        
        # Build schedule query with filters
        schedule_query = Schedule.query.filter_by(is_active=True)
        if day_filter:
            schedule_query = schedule_query.filter(Schedule.day_of_week == int(day_filter))
        if shift_filter:
            schedule_query = schedule_query.filter(Schedule.shift == shift_filter)
        
        filtered_schedules = schedule_query.all()
        
        # Create Excel file similar to export_excel but with filtered data
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Salas Filtradas"
        
        # Headers
        headers = ['ID', 'Nome', 'Capacidade', 'Bloco', 'Tem Computadores', 'Softwares', 'Descrição']
        for col, header in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col)
            if cell:
                cell.value = header
                try:
                    from openpyxl.styles import Font, PatternFill, Alignment
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                    cell.alignment = Alignment(horizontal='center')
                except ImportError:
                    pass  # Styles not available
        
        # Data
        for row, classroom in enumerate(filtered_classrooms, 2):
            ws1.cell(row=row, column=1).value = classroom.id
            ws1.cell(row=row, column=2).value = classroom.name
            ws1.cell(row=row, column=3).value = classroom.capacity
            ws1.cell(row=row, column=4).value = classroom.block
            ws1.cell(row=row, column=5).value = 'Sim' if classroom.has_computers else 'Não'
            ws1.cell(row=row, column=6).value = classroom.software
            ws1.cell(row=row, column=7).value = classroom.description
        
        # Auto-fit columns with safer approach
        try:
            for column_cells in ws1.columns:
                if column_cells and len(column_cells) > 0:
                    max_length = 0
                    column_letter = column_cells[0].column_letter
                    for cell in column_cells:
                        try:
                            if cell and cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    if ws1.column_dimensions and column_letter:
                        ws1.column_dimensions[column_letter].width = adjusted_width
        except Exception:
            pass  # Skip column sizing if there are issues
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f'relatorio_filtrado_{timestamp}.xlsx'
        
        response = make_response(output.getvalue())
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        
        return response
    except Exception as e:
        flash(f'Erro ao gerar Excel filtrado: {str(e)}', 'error')
        return redirect(url_for('dashboard'))

# Schedule Request Routes (for non-logged-in users)
@app.route('/request_schedule/<int:classroom_id>')
def request_schedule(classroom_id):
    """Show schedule request form for a specific classroom"""
    classroom = Classroom.query.get_or_404(classroom_id)
    return render_template('request_schedule.html', classroom=classroom)

@app.route('/submit_schedule_request', methods=['POST'])
def submit_schedule_request():
    """Process schedule request form submission"""
    try:
        # Get form data
        classroom_id = request.form.get('classroom_id')
        requester_name = request.form.get('requester_name', '').strip()
        requester_email = request.form.get('requester_email', '').strip()
        event_name = request.form.get('event_name', '').strip()
        description = request.form.get('description', '').strip()
        
        # Schedule details
        shift = request.form.get('shift', '').strip()
        start_time = request.form.get('start_time', '').strip()
        end_time = request.form.get('end_time', '').strip()
        
        # Check for bulk request (multiple dates)
        is_bulk_request = request.form.get('is_bulk_request') == 'on'
        
        if is_bulk_request:
            # Process date range with weekday selection
            start_date_str = request.form.get('start_date_bulk', '').strip()
            end_date_str = request.form.get('end_date_bulk', '').strip()
            weekdays = request.form.getlist('weekdays[]')
            
            if not start_date_str or not end_date_str or not weekdays:
                flash('Erro: Por favor, selecione o período e os dias da semana.', 'error')
                return redirect(url_for('request_schedule', classroom_id=classroom_id))
            
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
                weekdays = [int(w) for w in weekdays]
            except ValueError:
                flash('Erro: Formato de data inválido.', 'error')
                return redirect(url_for('request_schedule', classroom_id=classroom_id))
            
            if start_date > end_date:
                flash('Erro: A data inicial deve ser anterior à data final.', 'error')
                return redirect(url_for('request_schedule', classroom_id=classroom_id))
            
            # Generate all dates based on period and selected weekdays
            generated_dates = []
            current_date = start_date
            while current_date <= end_date:
                if current_date.weekday() in weekdays:
                    generated_dates.append(current_date.strftime('%Y-%m-%d'))
                current_date += timedelta(days=1)
            
            if not generated_dates:
                flash('Erro: Nenhuma data válida foi gerada com os critérios selecionados.', 'error')
                return redirect(url_for('request_schedule', classroom_id=classroom_id))
            
            # Use the first date as primary date
            primary_date = datetime.strptime(generated_dates[0], '%Y-%m-%d').date()
            primary_day_of_week = primary_date.weekday()
            
            # Store other dates as JSON
            other_dates = json.dumps(generated_dates[1:]) if len(generated_dates) > 1 else None
        else:
            # Single date request
            requested_date_str = request.form.get('requested_date', '').strip()
            if not requested_date_str:
                flash('Erro: Data é obrigatória.', 'error')
                return redirect(url_for('request_schedule', classroom_id=classroom_id))
            
            try:
                primary_date = datetime.strptime(requested_date_str, '%Y-%m-%d').date()
                primary_day_of_week = primary_date.weekday()
                other_dates = ''
            except ValueError:
                flash('Erro: Data inválida.', 'error')
                return redirect(url_for('request_schedule', classroom_id=classroom_id))
        
        # Validate required fields
        if not all([requester_name, requester_email, event_name, description, shift, start_time, end_time]):
            flash('Erro: Todos os campos obrigatórios devem ser preenchidos.', 'error')
            return redirect(url_for('request_schedule', classroom_id=classroom_id))
        
        # Create schedule request
        schedule_request = ScheduleRequest(
            classroom_id=int(classroom_id),
            requester_name=requester_name,
            requester_email=requester_email,
            event_name=event_name,
            description=description,
            requested_date=primary_date,
            day_of_week=primary_day_of_week,
            shift=shift,
            start_time=start_time,
            end_time=end_time,
            additional_dates=other_dates
        )
        
        db.session.add(schedule_request)
        db.session.commit()
        
        flash('Solicitação enviada com sucesso! qualquer duvida procure a gestão', 'success')
        return redirect(url_for('classroom_detail', classroom_id=classroom_id))
        
    except Exception as e:
        import logging
        logging.error(f"Error submitting schedule request: {str(e)}")
        db.session.rollback()
        flash('Erro ao enviar solicitação. Tente novamente.', 'error')
        return redirect(url_for('request_schedule', classroom_id=classroom_id))

# Admin routes for managing schedule requests
@app.route('/admin/schedule_requests')
@require_admin_auth
def admin_schedule_requests():
    """Admin page to view and manage schedule requests"""
    status_filter = request.args.get('status', 'pending')
    
    query = ScheduleRequest.query
    if status_filter and status_filter != 'all':
        query = query.filter(ScheduleRequest.status == status_filter)
    
    schedule_requests = query.order_by(ScheduleRequest.created_at.desc()).all()
    
    return render_template('admin_schedule_requests.html', 
                         requests=schedule_requests, 
                         current_status=status_filter)

@app.route('/admin/schedule_request/<int:request_id>/action', methods=['POST'])
@require_admin_auth
def admin_schedule_request_action(request_id):
    """Admin action to approve or reject schedule requests"""
    try:
        schedule_request = ScheduleRequest.query.get_or_404(request_id)
        # Debug: log all form data received
        import logging
        logging.info(f"Form data received: {dict(request.form)}")
        logging.info(f"Request method: {request.method}")
        logging.info(f"Content type: {request.content_type}")
        
        action = request.form.get('action')
        admin_notes = request.form.get('admin_notes', '').strip()
        
        # Debug: log the action received
        logging.info(f"Schedule request action received: '{action}' for request {request_id}")
        
        if action == 'approve':
            # Create schedule entries for approved request
            try:
                # Parse additional dates if they exist
                dates_to_schedule = [schedule_request.requested_date]
                if schedule_request.additional_dates:
                    additional_dates = json.loads(schedule_request.additional_dates)
                    for date_str in additional_dates:
                        dates_to_schedule.append(datetime.strptime(date_str, '%Y-%m-%d').date())
                
                # Create schedule entries for each date
                for schedule_date in dates_to_schedule:
                    new_schedule = Schedule(
                        classroom_id=schedule_request.classroom_id,
                        day_of_week=schedule_date.weekday(),
                        shift=schedule_request.shift,
                        course_name=schedule_request.event_name,
                        instructor=schedule_request.requester_name,
                        start_time=schedule_request.start_time,
                        end_time=schedule_request.end_time,
                        start_date=schedule_date,
                        end_date=schedule_date,  # Single day event
                        is_active=True
                    )
                    db.session.add(new_schedule)
                
                # Update request status before final commit
                schedule_request.status = 'approved'
                schedule_request.admin_notes = admin_notes
                schedule_request.reviewed_at = get_brazil_time()
                schedule_request.reviewed_by = 'Admin'
                
                # Commit all changes together
                db.session.commit()
                flash(f'Solicitação aprovada! {len(dates_to_schedule)} horário(s) adicionado(s) ao sistema.', 'success')
                return redirect(url_for('admin_schedule_requests'))
                
            except Exception as e:
                import logging
                logging.error(f"Error creating schedules from approved request: {str(e)}")
                logging.error(f"Request data: {schedule_request.__dict__}")
                db.session.rollback()
                flash(f'Erro ao criar horários no sistema: {str(e)}', 'error')
                return redirect(url_for('admin_schedule_requests'))
                
        elif action == 'reject':
            schedule_request.status = 'rejected'
            schedule_request.admin_notes = admin_notes
            schedule_request.reviewed_at = get_brazil_time()
            schedule_request.reviewed_by = 'Admin'
            db.session.commit()
            flash('Solicitação rejeitada.', 'info')
        else:
            flash('Ação inválida.', 'error')
            return redirect(url_for('admin_schedule_requests'))
        
    except Exception as e:
        import logging
        logging.error(f"Error processing schedule request action: {str(e)}")
        db.session.rollback()
        flash('Erro ao processar solicitação.', 'error')
    
    return redirect(url_for('admin_schedule_requests'))

# Template filters for proper data formatting
@app.template_filter('from_json')
def from_json(value):
    """Convert JSON string to Python object"""
    if not value:
        return []
    try:
        return json.loads(value)
    except (json.JSONDecodeError, TypeError):
        return []

@app.template_filter('dateformat')
def dateformat(value):
    """Format date string to Brazilian format"""
    try:
        if isinstance(value, str):
            date_obj = datetime.strptime(value, '%Y-%m-%d').date()
            return date_obj.strftime('%d/%m/%Y')
        return value
    except (ValueError, TypeError):
        return value

@app.route('/api/virtual-assistant', methods=['POST'])
def virtual_assistant():
    """Virtual Assistant endpoint for answering questions about classrooms and schedules"""
    try:
        data = request.get_json()
        user_message = data.get('message', '').strip().lower()
        
        if not user_message:
            return jsonify({'error': 'Mensagem não pode estar vazia'}), 400
            
        # Get current time for availability checks
        current_time = get_brazil_time()
        current_date = current_time.date()
        current_hour = current_time.hour
        current_weekday = current_time.weekday()  # 0=Monday, 6=Sunday
        
        # Get all classrooms and schedules
        classrooms = Classroom.query.all()
        schedules = Schedule.query.filter_by(is_active=True).filter(
            db.or_(
                Schedule.end_date == None,
                Schedule.end_date >= current_date
            )
        ).all()
        
        # Prepare response based on user question
        response = process_user_question(user_message, classrooms, schedules, current_time, current_date, current_hour, current_weekday)
        
        return jsonify({'response': response})
        
    except Exception as e:
        import logging
        logging.error(f"Error in virtual assistant: {str(e)}")
        return jsonify({'error': 'Erro interno do servidor. Tente novamente.'}), 500

def get_time_greeting(hour):
    """Return contextual greeting based on time of day"""
    if 5 <= hour < 12:
        return "Bom dia! ☀️"
    elif 12 <= hour < 18:
        return "Boa tarde! 🌤️"
    elif 18 <= hour < 22:
        return "Boa noite! 🌆"
    else:
        return "Oi! 🌙"

def process_user_question(user_message, classrooms, schedules, current_time, current_date, current_hour, current_weekday):
    """Process user question and return appropriate response - Always finds something useful to say!"""
    
    # Input safety
    if not user_message:
        return get_general_help_response()
    if not classrooms:
        return "🏫 Parece que ainda não temos salas cadastradas. Entre em contato com a secretaria para mais informações!"
    
    # Ensure safe data types
    schedules = schedules or []
    user_message_lower = user_message.lower()
    
    # Smart keyword detection with score system
    keyword_scores = {
        'availability': 0,
        'software': 0,
        'capacity': 0,
        'location': 0,
        'schedule': 0,
        'help': 0,
        'contact': 0,
        'about': 0,
        'analytics': 0
    }
    
    # Enhanced keyword matching with scoring
    availability_keywords = [
        'disponível', 'disponivel', 'livre', 'vaga', 'vazio', 'agora', 'now', 'aberta', 'ocupada', 'ocupado',
        'tem sala', 'preciso de sala', 'sala livre', 'sala vaga', 'reservar', 'usar sala', 'acesso', 'status'
    ]
    
    software_keywords = [
        'software', 'programa', 'aplicativo', 'aplicação', 'ferramenta', 'sistema',
        'unity', 'unreal', 'blender', 'visual studio', 'git', 'docker', 'office',
        'ide', 'editor', 'desenvolvimento', 'programação', 'programacao', 'código', 'codigo',
        'game', 'jogo', 'jogos', 'engine', '3d', 'modelagem', 'animação', 'animacao', 'computador'
    ]
    
    capacity_keywords = [
        'capacidade', 'quantas pessoas', 'quantos alunos', 'tamanho', 'lugares', 'assentos',
        'cabem', 'comporta', 'máximo', 'maximo', 'lotação', 'lotacao', 'turma', 'grupo', 'pessoal'
    ]
    
    location_keywords = [
        'onde', 'localização', 'localizacao', 'bloco', 'andar', 'fica', 'encontrar',
        'endereço', 'endereco', 'caminho', 'direção', 'direcao', 'mapa', 'local', 'chegar'
    ]
    
    schedule_keywords = [
        'horário', 'horario', 'aula', 'curso', 'quando', 'que horas', 'período', 'periodo',
        'manhã', 'manha', 'tarde', 'noite', 'segunda', 'terça', 'terca', 'quarta', 
        'quinta', 'sexta', 'sábado', 'sabado', 'domingo', 'funcionamento', 'aberto', 'programação'
    ]
    
    help_keywords = [
        'ajuda', 'help', 'como', 'o que', 'opções', 'opcoes', 'menu', 'comandos',
        'posso', 'consegue', 'sabe', 'funciona', 'usar', 'que você faz', 'oi', 'ola', 'olá'
    ]
    
    contact_keywords = [
        'contato', 'telefone', 'email', 'whatsapp', 'falar', 'secretaria', 'administração', 'administracao'
    ]
    
    about_keywords = [
        'senai', 'escola', 'instituição', 'instituicao', 'sobre', 'história', 'historia', 'morvan', 'figueiredo'
    ]
    
    analytics_keywords = [
        'análise', 'analise', 'tendência', 'tendencia', 'estatística', 'estatistica', 
        'padrão', 'padrao', 'histórico', 'historico', 'uso', 'ocupação', 'ocupacao',
        'relatório', 'relatorio', 'insights', 'dados', 'métricas', 'metricas', 'total', 'quantas'
    ]
    
    # Calculate scores for each intent
    keyword_sets = {
        'availability': availability_keywords,
        'software': software_keywords,
        'capacity': capacity_keywords,
        'location': location_keywords,
        'schedule': schedule_keywords,
        'help': help_keywords,
        'contact': contact_keywords,
        'about': about_keywords,
        'analytics': analytics_keywords
    }
    
    for intent, keywords in keyword_sets.items():
        for keyword in keywords:
            if keyword in user_message_lower:
                keyword_scores[intent] += 1
                # Give extra points for exact matches
                if keyword == user_message_lower.strip():
                    keyword_scores[intent] += 2
    
    # Find the highest scoring intent
    best_intent = max(keyword_scores, key=keyword_scores.get)
    best_score = keyword_scores[best_intent]
    
    # If we have a clear winner, use it
    if best_score > 0:
        if best_intent == 'availability':
            return get_available_rooms_now_smart(classrooms, schedules, current_time, current_date, current_hour, current_weekday)
        elif best_intent == 'software':
            return get_rooms_by_software_smart(user_message, classrooms)
        elif best_intent == 'capacity':
            return get_rooms_capacity_info_smart(classrooms)
        elif best_intent == 'location':
            return get_rooms_location_info_smart(classrooms)
        elif best_intent == 'schedule':
            return get_schedule_info_smart(classrooms, schedules)
        elif best_intent == 'contact':
            return get_contact_info()
        elif best_intent == 'about':
            return get_about_senai_info()
        elif best_intent == 'help':
            return get_general_help_response()
        elif best_intent == 'analytics':
            return get_analytics_and_trends_smart(classrooms, schedules, current_time)
    
    # If no clear intent, provide intelligent fallback
    return get_emergency_helpful_response(user_message, classrooms)

def get_available_rooms_now_smart(classrooms, schedules, current_time, current_date, current_hour, current_weekday):
    """Return information about currently available rooms with real-time database analysis"""
    try:
        from models import Classroom, Schedule
        from app import db
        from datetime import datetime, date, timedelta
        import pytz
        
        # Get São Paulo timezone
        sp_tz = pytz.timezone('America/Sao_Paulo')
        current_sp_time = datetime.now(sp_tz)
        current_date_sp = current_sp_time.date()
        current_hour_sp = current_sp_time.hour
        current_weekday_sp = current_sp_time.weekday()  # 0=Monday, 6=Sunday
        
        # Query real data from database
        all_classrooms = db.session.query(Classroom).all()
        active_schedules = db.session.query(Schedule).filter(
            Schedule.start_date <= current_date_sp,
            Schedule.end_date >= current_date_sp,
            Schedule.day_of_week == current_weekday_sp,
            Schedule.start_time <= str(current_hour_sp),
            Schedule.end_time > str(current_hour_sp)
        ).all()
        
        # Analyze real-time data
        available_rooms = []
        occupied_rooms = []
        
        for classroom in all_classrooms:
            is_occupied = False
            current_schedule = None
            
            # Check if classroom is currently occupied based on real schedule data
            for schedule in active_schedules:
                if schedule.classroom_id == classroom.id:
                    is_occupied = True
                    current_schedule = schedule
                    break
            
            if not is_occupied:
                available_rooms.append(classroom)
            else:
                occupied_rooms.append((classroom, current_schedule))
        
        # Get usage statistics for intelligent insights
        total_schedules_today = db.session.query(Schedule).filter(
            Schedule.start_date <= current_date_sp,
            Schedule.end_date >= current_date_sp,
            Schedule.day_of_week == current_weekday_sp
        ).count()
        
        # Get upcoming availability
        next_available = {}
        for room, schedule in occupied_rooms:
            if schedule:
                next_schedules = db.session.query(Schedule).filter(
                    Schedule.classroom_id == room.id,
                    Schedule.start_date <= current_date_sp,
                    Schedule.end_date >= current_date_sp,
                    Schedule.day_of_week == current_weekday_sp,
                    Schedule.start_time > str(current_hour_sp)
                ).order_by(Schedule.start_time).first()
                
                if next_schedules:
                    next_available[room.id] = next_schedules.start_time
                else:
                    next_available[room.id] = schedule.end_time
        
        # Generate intelligent, real-time analysis response
        time_greeting = get_time_greeting(current_hour_sp)
        
        if available_rooms:
            # Calculate availability percentage
            availability_percent = (len(available_rooms) / len(all_classrooms)) * 100
            
            response = f"{time_greeting} 😊\n\n"
            response += f"🟢 **Análise em tempo real - {len(available_rooms)} de {len(all_classrooms)} salas disponíveis ({availability_percent:.0f}% de disponibilidade):**\n\n"
            
            # Sort by capacity and usage patterns for smart recommendations
            available_rooms.sort(key=lambda x: x.capacity, reverse=True)
            
            for i, room in enumerate(available_rooms):
                # Check if room has schedules later today
                has_later_schedule = any(room.id in next_available.values() for room in available_rooms)
                
                response += f"{'🏆' if i == 0 else '⭐' if room.capacity >= 30 else '•'} **{room.name}** ({room.block})\n"
                response += f"  💺 {room.capacity} pessoas"
                if room.has_computers:
                    response += " | 💻 {0} computadores".format(room.capacity if room.has_computers else "Sem")
                response += "\n"
                
                if room.software:
                    response += f"  🛠️ Software: {room.software}\n"
                    
                # Add smart insights about room availability
                if room.id in next_available:
                    next_time = next_available[room.id]
                    response += f"  ⚠️ Ocupada às {next_time:02d}:00\n"
                else:
                    response += f"  ✅ Livre o resto do dia\n"
                    
                if room.description:
                    response += f"  📝 {room.description}\n"
                response += "\n"
            
            # Add intelligent insights
            response += f"📊 **Insights do Sistema:**\n"
            response += f"• {total_schedules_today} aulas programadas hoje\n"
            response += f"• Taxa de ocupação atual: {100-availability_percent:.0f}%\n"
            response += f"• Melhor horário: Salas mais disponíveis pela manhã\n\n"
            
            response += "🎯 **Recomendação inteligente:** Use a primeira sala da lista para maior flexibilidade!\n"
            response += "💬 Reservas: Solicite no sistema"
            
        else:
            response = f"{time_greeting} 😅\n\n"
            response += f"🔴 **Análise: Todas as {len(all_classrooms)} salas estão ocupadas ({current_sp_time.strftime('%H:%M')})**\n\n"
            
            if occupied_rooms:
                response += "📚 **Atividades em andamento (dados em tempo real):**\n\n"
                for room, schedule in occupied_rooms[:4]:  # Show first 4
                    response += f"• **{room.name}** ({room.block})"
                    if schedule and hasattr(schedule, 'course_name'):
                        response += f" - {schedule.course_name}"
                        if hasattr(schedule, 'end_time'):
                            response += f" (até {schedule.end_time:02d}:00)"
                    response += "\n"
                
                if len(occupied_rooms) > 4:
                    response += f"... e mais {len(occupied_rooms) - 4} salas ocupadas\n"
                
                # Show when rooms will be free
                response += "\n⏰ **Próximas liberações:**\n"
                liberation_times = {}
                for room, schedule in occupied_rooms:
                    if schedule and hasattr(schedule, 'end_time'):
                        end_time = schedule.end_time
                        if end_time not in liberation_times:
                            liberation_times[end_time] = []
                        liberation_times[end_time].append(room.name)
                
                for time, rooms in sorted(liberation_times.items()):
                    response += f"• {time:02d}:00 - {', '.join(rooms[:2])}"
                    if len(rooms) > 2:
                        response += f" (+{len(rooms)-2} outras)"
                    response += "\n"
            
            response += f"\n📊 **Estatística do dia:** {total_schedules_today} atividades programadas\n"
            response += "🔄 **Tente em alguns minutos ou pergunte sobre horários específicos!**"
        
        return response
        
    except Exception as e:
        import logging
        logging.error(f"Error in get_available_rooms_now_smart: {str(e)}")
        # Never give up - provide basic info at least
        return get_basic_classroom_info(classrooms)

def get_rooms_by_software_smart(user_message, classrooms):
    """Return rooms that have specific software with real-time database analysis"""
    try:
        from models import Classroom, Schedule
        from app import db
        from datetime import datetime, date
        import pytz
        
        # Get real-time data from database
        sp_tz = pytz.timezone('America/Sao_Paulo')
        current_sp_time = datetime.now(sp_tz)
        current_date_sp = current_sp_time.date()
        current_hour_sp = current_sp_time.hour
        current_weekday_sp = current_sp_time.weekday()
        
        # Query all classrooms with real data
        all_classrooms = db.session.query(Classroom).all()
        
        # Enhanced software detection with intelligent matching
        software_keywords = {
            'unity': ['unity', 'engine unity', 'game engine'],
            'unreal': ['unreal', 'unreal engine', 'ue4', 'ue5'],
            'blender': ['blender', '3d', 'modelagem', 'animação', 'animacao'],
            'visual studio': ['visual studio', 'vs', 'ide', 'desenvolvimento', 'programação', 'programacao'],
            'git': ['git', 'versionamento', 'controle de versão'],
            'docker': ['docker', 'container', 'containerização'],
            'office': ['office', 'word', 'excel', 'powerpoint', 'escritório'],
            'banco de dados': ['banco', 'database', 'bd', 'sql', 'mysql', 'postgresql'],
            'jogos': ['jogo', 'jogos', 'game', 'games', 'desenvolvimento de jogos']
        }
        
        # Find software mentioned in user message
        mentioned_software = []
        for software_type, keywords in software_keywords.items():
            if any(keyword.lower() in user_message.lower() for keyword in keywords):
                mentioned_software.append(software_type)
        
        # Analyze real database data for software and availability
        matching_rooms = []
        all_software_rooms = []
        
        for classroom in all_classrooms:
            if classroom.software:
                all_software_rooms.append(classroom)
                software_lower = classroom.software.lower()
                
                # Check availability in real-time
                try:
                    is_available_now = not db.session.query(Schedule).filter(
                        Schedule.classroom_id == classroom.id,
                        Schedule.start_date <= current_date_sp,
                        Schedule.end_date >= current_date_sp,
                        Schedule.day_of_week == current_weekday_sp,
                        Schedule.start_time.cast(db.Integer) <= current_hour_sp,
                        Schedule.end_time.cast(db.Integer) > current_hour_sp
                    ).first()
                except:
                    # Fallback if casting fails
                    is_available_now = True
                
                # Check if any mentioned software is in this classroom
                for software_type in mentioned_software:
                    keywords = software_keywords[software_type]
                    if any(keyword.lower() in software_lower for keyword in keywords):
                        matching_rooms.append((classroom, software_type, is_available_now))
                        break
                        
        # Get usage statistics for this software
        software_usage_stats = {}
        for software_type in mentioned_software:
            total_rooms_with_software = sum(1 for room in all_classrooms 
                                          if room.software and any(kw.lower() in room.software.lower() 
                                          for kw in software_keywords[software_type]))
            available_now = sum(1 for room, _, available in matching_rooms if available)
            software_usage_stats[software_type] = {
                'total': total_rooms_with_software,
                'available_now': available_now
            }
        
        # Generate intelligent response
        if mentioned_software:
            software_list = ", ".join(mentioned_software).title()
            response = f"🔍 **{software_list}:**\n\n"
            
            if matching_rooms:
                response += f"🎯 **Salas com {software_list}:**\n\n"
                
                # Sort by availability first, then by capacity
                matching_rooms.sort(key=lambda x: (not x[2], -x[0].capacity))
                
                for i, (room, software_type, is_available) in enumerate(matching_rooms):
                    # Dynamic emoji based on availability and capacity
                    if is_available and room.capacity >= 30:
                        emoji = "🏆"  # Best option
                    elif is_available:
                        emoji = "✅"  # Available
                    elif room.capacity >= 30:
                        emoji = "⭐"  # Large but occupied
                    else:
                        emoji = "🔴"  # Occupied
                    
                    response += f"{emoji} **{room.name}** ({room.block}) "
                    response += f"{'🟢 DISPONÍVEL' if is_available else '🔴 OCUPADA'}\n"
                    response += f"  💺 {room.capacity} pessoas"
                    if room.has_computers:
                        response += " | 💻 Com computadores"
                    response += f"\n  🛠️ {room.software}\n"
                    
                    # Add real-time insights
                    if not is_available:
                        # Check when it will be free
                        try:
                            next_free = db.session.query(Schedule).filter(
                                Schedule.classroom_id == room.id,
                                Schedule.start_date <= current_date_sp,
                                Schedule.end_date >= current_date_sp,
                                Schedule.day_of_week == current_weekday_sp,
                                Schedule.start_time.cast(db.Integer) <= current_hour_sp,
                                Schedule.end_time.cast(db.Integer) > current_hour_sp
                            ).first()
                        except:
                            next_free = None
                        
                        if next_free:
                            response += f"  ⏰ Livre às {next_free.end_time:02d}:00\n"
                    
                    if room.description:
                        response += f"  📝 {room.description}\n"
                    response += "\n"
                
                # Add intelligent statistics
                for software_type, stats in software_usage_stats.items():
                    response += f"📊 **{software_type.title()}:** {stats['available_now']}/{stats['total']} salas disponíveis agora\n"
                
                response += "\n💡 **Dica:** Salas com ✅ estão livres para uso imediato!\n"
            else:
                response += f"😅 **Hmm... não encontrei salas específicas com {software_list}.**\n\n"
                response += "Mas deixe-me mostrar todas as opções disponíveis:\n\n"
        else:
            response = "💻 **Que software você está procurando?** Aqui estão todas nossas opções! 😊\n\n"
        
        # Show all software rooms if no specific match or no software mentioned
        if not matching_rooms or not mentioned_software:
            if all_software_rooms:
                response += "📋 **Todas as salas com software disponível:**\n\n"
                
                # Group by software type for better organization
                software_groups = {}
                for room in all_software_rooms:
                    key_software = room.software.split(',')[0].strip() if ',' in room.software else room.software
                    if key_software not in software_groups:
                        software_groups[key_software] = []
                    software_groups[key_software].append(room)
                
                for software, rooms in software_groups.items():
                    response += f"🔧 **{software}:**\n"
                    for room in rooms:
                        response += f"  • **{room.name}** ({room.block}) - {room.capacity} pessoas\n"
                    response += "\n"
            else:
                response += "😅 **Ops! Parece que não temos informações de software cadastradas ainda.**\n"
                response += "Entre em contato com a secretaria para mais detalhes! 📞"
        
        response += f"\n🤖 **Análise concluída às {current_sp_time.strftime('%H:%M')}!** Pergunte sobre disponibilidade, localização ou qualquer outra dúvida!"
        
        return response
        
    except Exception as e:
        import logging
        logging.error(f"Error in get_rooms_by_software_smart: {str(e)}")
        # Always show available software options
        return get_all_software_options(classrooms)

def get_rooms_capacity_info_smart(classrooms):
    """Return information about room capacities with real-time database analysis"""
    try:
        from models import Classroom, Schedule
        from app import db
        from datetime import datetime, date, timedelta
        import pytz
        
        # Get real-time data from database
        sp_tz = pytz.timezone('America/Sao_Paulo')
        current_sp_time = datetime.now(sp_tz)
        current_date_sp = current_sp_time.date()
        current_hour_sp = current_sp_time.hour
        current_weekday_sp = current_sp_time.weekday()
        
        # Query all classrooms with real data
        all_classrooms = db.session.query(Classroom).all()
        
        if not all_classrooms:
            return "😅 Ops! Não encontrei informações sobre as salas. Tente novamente! 🤗"
        
        # Organize rooms by capacity ranges with availability analysis
        small_rooms = []  # 1-20 people
        medium_rooms = []  # 21-35 people
        large_rooms = []  # 36+ people
        
        for room in all_classrooms:
            if hasattr(room, 'capacity') and room.capacity:
                # Check current availability
                try:
                    is_available_now = not db.session.query(Schedule).filter(
                        Schedule.classroom_id == room.id,
                        Schedule.start_date <= current_date_sp,
                        Schedule.end_date >= current_date_sp,
                        Schedule.day_of_week == current_weekday_sp,
                        Schedule.start_time.cast(db.Integer) <= current_hour_sp,
                        Schedule.end_time.cast(db.Integer) > current_hour_sp
                    ).first()
                except:
                    is_available_now = True
                
                # Calculate weekly usage (how many hours per week this room is scheduled)
                weekly_usage = db.session.query(Schedule).filter(
                    Schedule.classroom_id == room.id,
                    Schedule.start_date <= current_date_sp,
                    Schedule.end_date >= current_date_sp
                ).count()
                
                room_data = (room, is_available_now, weekly_usage)
                
                if room.capacity <= 20:
                    small_rooms.append(room_data)
                elif room.capacity <= 35:
                    medium_rooms.append(room_data)
                else:
                    large_rooms.append(room_data)
        
        # Sort each category by availability first, then capacity
        small_rooms.sort(key=lambda x: (not x[1], x[0].capacity))
        medium_rooms.sort(key=lambda x: (not x[1], x[0].capacity))
        large_rooms.sort(key=lambda x: (not x[1], -x[0].capacity))
        
        response = f"👥 **Capacidade das Salas ({current_sp_time.strftime('%H:%M')}):**\n\n"
        
        # Calculate real-time statistics
        total_rooms = len(all_classrooms)
        available_now = sum(1 for category in [small_rooms, medium_rooms, large_rooms] 
                           for room, available, _ in category if available)
        total_capacity = sum(room.capacity for room in all_classrooms if hasattr(room, 'capacity') and room.capacity)
        
        if large_rooms:
            available_large = sum(1 for _, available, _ in large_rooms if available)
            response += f"🏢 **Salas Grandes (35+ pessoas) - {available_large}/{len(large_rooms)} disponíveis:**\n"
            for room, is_available, weekly_usage in large_rooms:
                status_emoji = "🟢" if is_available else "🔴"
                usage_level = "🔥 Alta" if weekly_usage > 10 else "📊 Média" if weekly_usage > 5 else "💤 Baixa"
                
                response += f"  {status_emoji} **{room.name}** ({room.block}) - **{room.capacity} pessoas**"
                if room.has_computers:
                    response += " | 💻 Com computadores"
                response += f"\n    📈 Uso semanal: {usage_level} ({weekly_usage} horários)\n"
                if room.software:
                    response += f"    🛠️ {room.software}\n"
            response += "\n"
        
        if medium_rooms:
            available_medium = sum(1 for _, available, _ in medium_rooms if available)
            response += f"🏤 **Salas Médias (21-35 pessoas) - {available_medium}/{len(medium_rooms)} disponíveis:**\n"
            for room, is_available, weekly_usage in medium_rooms:
                status_emoji = "🟢" if is_available else "🔴"
                usage_level = "🔥 Alta" if weekly_usage > 10 else "📊 Média" if weekly_usage > 5 else "💤 Baixa"
                
                response += f"  {status_emoji} **{room.name}** ({room.block}) - **{room.capacity} pessoas**"
                if room.has_computers:
                    response += " | 💻 Com computadores"
                response += f"\n    📈 Uso semanal: {usage_level} ({weekly_usage} horários)\n"
                if room.software:
                    response += f"    🛠️ {room.software}\n"
            response += "\n"
        
        if small_rooms:
            available_small = sum(1 for _, available, _ in small_rooms if available)
            response += f"🏠 **Salas Menores (até 20 pessoas) - {available_small}/{len(small_rooms)} disponíveis:**\n"
            for room, is_available, weekly_usage in small_rooms:
                status_emoji = "🟢" if is_available else "🔴"
                usage_level = "🔥 Alta" if weekly_usage > 10 else "📊 Média" if weekly_usage > 5 else "💤 Baixa"
                
                response += f"  {status_emoji} **{room.name}** ({room.block}) - **{room.capacity} pessoas**"
                if room.has_computers:
                    response += " | 💻 Com computadores"
                response += f"\n    📈 Uso semanal: {usage_level} ({weekly_usage} horários)\n"
                if room.software:
                    response += f"    🛠️ {room.software}\n"
            response += "\n"
        
        # Add intelligent analytics
        avg_capacity = total_capacity / total_rooms if total_rooms else 0
        utilization_rate = ((total_rooms - available_now) / total_rooms) * 100 if total_rooms else 0
        
        response += f"📊 **Análise Inteligente do Sistema:**\n"
        response += f"• **Disponibilidade atual:** {available_now}/{total_rooms} salas ({(available_now/total_rooms)*100:.0f}%)\n"
        response += f"• **Taxa de ocupação:** {utilization_rate:.0f}%\n"
        response += f"• **Capacidade total:** {total_capacity} pessoas\n"
        response += f"• **Capacidade média:** {avg_capacity:.0f} pessoas/sala\n"
        response += f"• **Eficiência do espaço:** {'🟢 Ótima' if utilization_rate < 70 else '🟡 Boa' if utilization_rate < 85 else '🔴 Alta demanda'}\n\n"
        
        # Smart recommendations based on real data
        if available_now > 0:
            response += "🎯 **Recomendações inteligentes:**\n"
            response += "• Salas com 🟢 estão livres para uso imediato\n"
            response += "• Salas com uso 💤 Baixo são ideais para reservas futuras\n"
        else:
            response += "⚠️ **Alta demanda detectada!** Considere agendar com antecedência\n"
        
        response += f"\n💡 **Análise atualizada a cada consulta em tempo real!** 🔄"
        
        return response
        
    except Exception as e:
        import logging
        logging.error(f"Error in get_rooms_capacity_info_smart: {str(e)}")
        # Always show basic capacity info
        return get_basic_capacity_info(classrooms)
    
    # Sort rooms by capacity
    sorted_rooms = sorted(classrooms, key=lambda x: x.capacity, reverse=True)
    
    for room in sorted_rooms:
        response += f"• **{room.name}** ({room.block}) - {room.capacity} pessoas\n"
    
    total_capacity = sum(room.capacity for room in classrooms)
    response += f"\n📊 **Capacidade total:** {total_capacity} pessoas"
    
    return response

def get_rooms_location_info(classrooms):
    """Return information about room locations"""
    response = "📍 **Localização das Salas:**\n\n"
    
    # Group by block
    blocks = {}
    for room in classrooms:
        if room.block not in blocks:
            blocks[room.block] = []
        blocks[room.block].append(room)
    
    for block, rooms in blocks.items():
        response += f"**{block}:**\n"
        for room in rooms:
            response += f"  • {room.name} (capacidade: {room.capacity})\n"
        response += "\n"
    
    return response

def get_schedule_info(classrooms, schedules):
    """Return general schedule information"""
    response = "📅 **Informações sobre Horários:**\n\n"
    
    weekdays = ['Segunda', 'Terça', 'Quarta', 'Quinta', 'Sexta', 'Sábado', 'Domingo']
    
    active_schedules = [s for s in schedules if s.is_active]
    
    if active_schedules:
        response += f"Existem **{len(active_schedules)} horários ativos** nas salas.\n\n"
        
        # Show schedules by day
        for day in range(7):  # 0=Monday, 6=Sunday
            day_schedules = [s for s in active_schedules if s.weekday == day]
            if day_schedules:
                response += f"**{weekdays[day]}:**\n"
                for schedule in day_schedules:
                    classroom = next((c for c in classrooms if c.id == schedule.classroom_id), None)
                    if classroom:
                        response += f"  • {classroom.name}: {schedule.start_time:02d}h - {schedule.end_time:02d}h ({schedule.course_name})\n"
                response += "\n"
    else:
        response += "Não há horários ativos no momento.\n"
    
    return response

def get_general_help_response():
    """Return general help information with personality"""
    return """🤖 **Oi! Sou seu assistente virtual do SENAI Morvan Figueiredo! 😊**

Estou aqui para te ajudar com tudo sobre nossas salas e laboratórios. Sou bem esperto e converso naturalmente - não precisa usar comandos específicos! 🗣️

**🎯 Exemplos do que posso fazer por você:**

**🏢 Sobre as Salas:**
💬 *"Preciso de uma sala para 25 pessoas com computadores"*
💬 *"Onde fica a Sala DEV?"*
💬 *"Que salas têm Unity para desenvolvimento de jogos?"*

**⚡ Disponibilidade em Tempo Real:**
💬 *"Que salas estão livres agora?"*
💬 *"A sala de jogos está ocupada?"*
💬 *"Quando o lab fica disponível?"*

**🛠️ Software e Tecnologia:**
💬 *"Preciso usar Blender para modelagem 3D"*
💬 *"Onde tem Visual Studio?"*
💬 *"Sala com banco de dados MySQL"*

**📊 Informações Gerais:**
💬 *"Como funciona o SENAI?"*
💬 *"Telefone para contato"*
💬 *"Horários de funcionamento"*

**💡 Minha especialidade:** Entendo linguagem natural! Não precisa usar comandos específicos - apenas me fale normalmente o que você precisa! 

**🚀 Exemplo:** Em vez de perguntar "salas capacidade", me pergunte *"Preciso de uma sala grande para apresentação"* que eu entendo perfeitamente! 😉

**🤝 Estou sempre aprendendo!** Se não entender alguma coisa, me explique de outra forma que eu vou me adaptar! 🧠✨"""

def get_analytics_and_trends(classrooms, schedules, current_time):
    """Return comprehensive analytics and trends from real database data"""
    try:
        from models import Classroom, Schedule, Incident
        from app import db
        from datetime import datetime, date, timedelta
        import pytz
        from collections import defaultdict
        
        # Get real-time data
        sp_tz = pytz.timezone('America/Sao_Paulo')
        current_sp_time = datetime.now(sp_tz)
        current_date_sp = current_sp_time.date()
        current_weekday_sp = current_sp_time.weekday()
        
        # Query comprehensive data
        all_classrooms = db.session.query(Classroom).all()
        all_schedules = db.session.query(Schedule).all()
        all_incidents = db.session.query(Incident).all()
        
        # Analytics calculations
        total_rooms = len(all_classrooms)
        total_schedules = len(all_schedules)
        total_incidents = len(all_incidents)
        
        response = f"📊 **Análise Completa do Sistema SENAI - {current_sp_time.strftime('%d/%m/%Y %H:%M')}**\n\n"
        
        # === OCCUPANCY ANALYSIS ===
        weekday_usage = defaultdict(int)
        hour_usage = defaultdict(int)
        room_popularity = defaultdict(int)
        
        for schedule in all_schedules:
            if hasattr(schedule, 'weekday') and hasattr(schedule, 'start_time'):
                weekday_usage[schedule.weekday] += 1
                hour_usage[schedule.start_time] += 1
                if hasattr(schedule, 'classroom_id'):
                    room_popularity[schedule.classroom_id] += 1
        
        # Current availability analysis
        currently_occupied = 0
        currently_available = 0
        
        for classroom in all_classrooms:
            is_occupied = db.session.query(Schedule).filter(
                Schedule.classroom_id == classroom.id,
                Schedule.start_date <= current_date_sp,
                Schedule.end_date >= current_date_sp,
                Schedule.day_of_week == current_weekday_sp,
                Schedule.start_time <= current_sp_time.hour,
                Schedule.end_time > current_sp_time.hour
            ).first()
            
            if is_occupied:
                currently_occupied += 1
            else:
                currently_available += 1
        
        # === REAL-TIME STATUS ===
        response += "🎯 **Status Atual do Sistema:**\n"
        response += f"• **Ocupação em tempo real:** {currently_occupied}/{total_rooms} salas ({(currently_occupied/total_rooms)*100:.1f}%)\n"
        response += f"• **Disponibilidade:** {currently_available} salas livres\n"
        response += f"• **Total de agendamentos:** {total_schedules} horários cadastrados\n"
        response += f"• **Incidentes registrados:** {total_incidents} ocorrências\n\n"
        
        # === USAGE PATTERNS ===
        response += "📈 **Padrões de Uso Inteligente:**\n"
        
        # Busiest day
        if weekday_usage:
            days = ['Segunda', 'Terça', 'Quarta', 'Quinta', 'Sexta', 'Sábado', 'Domingo']
            busiest_day_num = max(weekday_usage.keys(), key=lambda x: weekday_usage[x])
            busiest_day_count = weekday_usage[busiest_day_num]
            response += f"• **Dia mais movimentado:** {days[busiest_day_num]} ({busiest_day_count} agendamentos)\n"
        
        # Peak hours
        if hour_usage:
            peak_hour = max(hour_usage.keys(), key=lambda x: hour_usage[x])
            peak_count = hour_usage[peak_hour]
            response += f"• **Horário de pico:** {peak_hour:02d}:00 ({peak_count} salas em uso)\n"
        
        # Most popular rooms
        if room_popularity:
            most_used_room_id = max(room_popularity.keys(), key=lambda x: room_popularity[x])
            most_used_room = next((room for room in all_classrooms if room.id == most_used_room_id), None)
            if most_used_room:
                usage_count = room_popularity[most_used_room_id]
                response += f"• **Sala mais utilizada:** {most_used_room.name} ({usage_count} agendamentos)\n"
        
        response += "\n"
        
        # === CAPACITY OPTIMIZATION ===
        total_capacity = sum(room.capacity for room in all_classrooms if hasattr(room, 'capacity'))
        avg_capacity = total_capacity / total_rooms if total_rooms else 0
        
        response += "🏗️ **Otimização de Espaços:**\n"
        response += f"• **Capacidade total:** {total_capacity} pessoas simultaneamente\n"
        response += f"• **Capacidade média:** {avg_capacity:.0f} pessoas por sala\n"
        
        # Calculate efficiency
        peak_usage_percent = (currently_occupied / total_rooms) * 100 if total_rooms else 0
        if peak_usage_percent < 60:
            efficiency = "🟢 Eficiente - Boa disponibilidade"
        elif peak_usage_percent < 80:
            efficiency = "🟡 Moderada - Ocupação balanceada"
        else:
            efficiency = "🔴 Alta demanda - Considere expansão"
            
        response += f"• **Eficiência atual:** {efficiency}\n\n"
        
        # === TECHNOLOGY INSIGHTS ===
        rooms_with_computers = sum(1 for room in all_classrooms if hasattr(room, 'has_computers') and room.has_computers)
        rooms_with_software = sum(1 for room in all_classrooms if hasattr(room, 'software') and room.software)
        
        response += "💻 **Análise Tecnológica:**\n"
        response += f"• **Salas informatizadas:** {rooms_with_computers}/{total_rooms} ({(rooms_with_computers/total_rooms)*100:.0f}%)\n"
        response += f"• **Salas com software especializado:** {rooms_with_software}/{total_rooms}\n"
        
        # Software distribution
        software_count = defaultdict(int)
        for room in all_classrooms:
            if hasattr(room, 'software') and room.software:
                # Count main software types
                if 'unity' in room.software.lower():
                    software_count['Unity'] += 1
                if 'blender' in room.software.lower():
                    software_count['Blender'] += 1
                if 'visual studio' in room.software.lower():
                    software_count['Visual Studio'] += 1
        
        if software_count:
            response += "• **Software mais comum:** "
            top_software = max(software_count.keys(), key=lambda x: software_count[x])
            response += f"{top_software} ({software_count[top_software]} salas)\n"
        
        response += "\n"
        
        # === MAINTENANCE INSIGHTS ===
        if total_incidents > 0:
            response += "🔧 **Análise de Manutenção:**\n"
            response += f"• **Total de incidentes:** {total_incidents} registros\n"
            response += f"• **Média de incidentes:** {total_incidents/total_rooms:.1f} por sala\n"
            
            # Most problematic rooms
            incident_count = defaultdict(int)
            for incident in all_incidents:
                if hasattr(incident, 'classroom_id'):
                    incident_count[incident.classroom_id] += 1
            
            if incident_count:
                problematic_room_id = max(incident_count.keys(), key=lambda x: incident_count[x])
                problematic_room = next((room for room in all_classrooms if room.id == problematic_room_id), None)
                if problematic_room:
                    response += f"• **Sala que requer atenção:** {problematic_room.name} ({incident_count[problematic_room_id]} incidentes)\n"
            response += "\n"
        
        # === PREDICTIONS AND RECOMMENDATIONS ===
        response += "🔮 **Insights e Recomendações:**\n"
        
        if peak_usage_percent > 80:
            response += "• ⚠️ **Alta demanda detectada** - Considere otimizar horários\n"
        elif peak_usage_percent < 40:
            response += "• 💡 **Baixa ocupação** - Oportunidade para novos cursos\n"
        
        if currently_available > currently_occupied:
            response += "• ✅ **Boa disponibilidade** - Momento ideal para reservas\n"
        
        # Time-based recommendations
        current_hour = current_sp_time.hour
        if 8 <= current_hour <= 10:
            response += "• 🌅 **Período matutino** - Horário de menor demanda\n"
        elif 14 <= current_hour <= 16:
            response += "• 🌞 **Período vespertino** - Pico de atividades\n"
        elif 19 <= current_hour <= 21:
            response += "• 🌆 **Período noturno** - Alta demanda por cursos\n"
        
        response += f"\n🔄 **Análise atualizada automaticamente - Última atualização: {current_sp_time.strftime('%H:%M:%S')}**"
        response += f"\n💡 **Dados baseados em {total_schedules} agendamentos e {total_incidents} registros históricos**"
        
        return response
        
    except Exception as e:
        import logging
        logging.error(f"Error in get_analytics_and_trends: {str(e)}")
        # Always provide some useful analytics info
        return get_analytics_and_trends_smart(classrooms, schedules, current_time)

def get_contact_info():
    """Return contact information"""
    return """📞 **Informações de Contato - SENAI Morvan Figueiredo:**

🏢 **Endereço:**
SENAI "Morvan Figueiredo" - CFP 1.03
São Paulo, SP

📧 **Secretaria:**
Para informações sobre cursos, matrículas e horários, entre em contato com a secretaria da escola.

🕒 **Horário de Funcionamento:**
• Manhã: 7h às 12h
• Tarde: 13h às 18h  
• Noite: 19h às 22h

💻 **Sistema de Salas:**
Este assistente virtual pode ajudar com informações sobre:
• Disponibilidade das salas
• Software instalado
• Capacidades e localizações
• Horários de uso

Para dúvidas administrativas, procure a secretaria presencialmente! 😊"""

def get_about_senai_info():
    """Return information about SENAI"""
    return """🏫 **Sobre o SENAI Morvan Figueiredo:**

📚 **O que é o SENAI:**
O Serviço Nacional de Aprendizagem Industrial (SENAI) é a principal rede de educação profissional do país, oferecendo cursos técnicos e de qualificação profissional.

🎯 **SENAI "Morvan Figueiredo" - CFP 1.03:**
• Foco em **Tecnologia da Informação** e **Desenvolvimento de Jogos**
• Laboratórios equipados com software profissional
• Cursos técnicos e de qualificação
• Formação prática para o mercado de trabalho

💻 **Nossas Salas:**
• **Laboratório de Jogos Digitais** - Unity, Unreal Engine, Blender
• **Sala DEV** - Visual Studio, Git, Docker
• **Salas 208 e 202** - IDE, Banco de dados, Office

🌟 **Missão:**
Formar profissionais qualificados para a indústria, contribuindo para o desenvolvimento tecnológico e econômico do país.

Quer saber mais sobre as salas e horários? Use os botões de sugestão! 🚀"""

def get_smart_fallback_response(user_message, classrooms, schedules, current_time):
    """Intelligent fallback response with context analysis"""
    try:
        # Analyze the user message for context clues
        message_lower = user_message.lower()
        
        # Smart context detection
        context_hints = []
        
        # Detect mentions of specific rooms
        mentioned_rooms = []
        for classroom in classrooms:
            if classroom.name.lower() in message_lower:
                mentioned_rooms.append(classroom)
        
        # Detect time-related queries
        time_keywords = ['quando', 'que horas', 'horário', 'horario', 'tempo', 'duração', 'duracao']
        is_time_query = any(keyword in message_lower for keyword in time_keywords)
        
        # Detect frustration or confusion
        confused_keywords = ['não entendi', 'nao entendi', 'confuso', 'help', 'socorro', 'não sei', 'nao sei']
        is_confused = any(keyword in message_lower for keyword in confused_keywords)
        
        # Detect greeting or casual conversation
        greeting_keywords = ['oi', 'olá', 'ola', 'bom dia', 'boa tarde', 'boa noite', 'tchau', 'obrigado', 'obrigada', 'valeu']
        is_greeting = any(keyword in message_lower for keyword in greeting_keywords)
        
        # Generate contextual response
        if is_confused:
            return """😅 **Vejo que você está com dúvida! Sem problemas, estou aqui para ajudar! 🤗**

Vou te dar algumas dicas para conversarmos melhor:

**🎯 Tente perguntas assim:**
• *"Preciso de uma sala para 20 pessoas"*
• *"Onde fica a Sala DEV?"*
• *"Que salas estão livres agora?"*
• *"Quais salas têm Unity?"*

**💡 Dica especial:** Fale comigo como se fosse um amigo! Não precisa usar linguagem técnica. 

**Exemplo:** Em vez de *"consultar disponibilidade salas"*, me pergunte *"tem alguma sala livre agora?"* 😊

**🤝 Vamos tentar de novo?** Me conte o que você precisa de uma forma simples e natural! Estou aqui para te ajudar! ✨"""
        
        elif is_greeting:
            time_greeting = get_time_greeting(current_time.hour)
            return f"""🤗 **{time_greeting}**

Que bom te ver por aqui! Sou o assistente virtual do SENAI Morvan Figueiredo e estou super animado para te ajudar! 😊

**🎯 Posso te ajudar com:**
• 🏢 Informações sobre salas e laboratórios
• ⏰ Disponibilidade em tempo real
• 💻 Software e equipamentos
• 📍 Localização e direções
• 📞 Contatos e horários

**💬 Como posso te ajudar hoje?** Pode me perguntar qualquer coisa sobre nossas instalações! 

Exemplo: *"Preciso de uma sala com computadores"* ou *"Onde fica o laboratório de jogos?"* 🚀"""
        
        elif mentioned_rooms:
            room = mentioned_rooms[0]
            return f"""🎯 **Vi que você mencionou a {room.name}! Aqui estão as informações:**

**📍 Localização:** {room.block}
**👥 Capacidade:** {room.capacity} pessoas
**💻 Computadores:** {'Sim' if room.has_computers else 'Não'}
{f"**🛠️ Software:** {room.software}" if room.software else ""}
{f"**📝 Descrição:** {room.description}" if room.description else ""}

**🤔 O que você gostaria de saber sobre esta sala?**
• Se está disponível agora?
• Como chegar até lá?
• Mais detalhes sobre os equipamentos?

É só me perguntar! 😊"""
        
        elif is_time_query:
            return """⏰ **Perguntas sobre horários? Posso te ajudar! 😊**

**📅 Posso te dizer:**
• Que salas estão livres agora
• Quando uma sala específica fica disponível
• Horários de funcionamento do SENAI
• Quando termina uma aula específica

**💬 Exemplos de como perguntar:**
• *"Que horas a Sala DEV fica livre?"*
• *"Até que horas funciona o SENAI?"*
• *"Quando termina a aula de jogos?"*
• *"Que salas estão disponíveis agora?"*

**🕒 Horário atual:** {current_time.strftime('%H:%M')}

**❓ Sobre que horário você gostaria de saber?**"""
        
        # Default intelligent response
        return f"""🤖 **Hmm... não tenho certeza do que você está procurando, mas vou te ajudar! 😊**

**🔍 Analisando sua mensagem:** *"{user_message}"*

**💡 Algumas sugestões baseadas no que você disse:**

**🏢 Se for sobre salas:**
• *"Que salas estão livres agora?"*
• *"Preciso de uma sala para X pessoas"*
• *"Onde fica a [nome da sala]?"*

**💻 Se for sobre software/equipamentos:**
• *"Quais salas têm [nome do software]?"*
• *"Preciso usar [programa específico]"*

**📞 Se for sobre contato/informações:**
• *"Como entrar em contato?"*
• *"Horário de funcionamento"*
• *"Sobre o SENAI"*

**🤝 Reformule sua pergunta de forma mais específica e eu vou te dar uma resposta perfeita! 🎯**

**⏰ Horário atual:** {current_time.strftime('%H:%M')} - {"📅 " + current_time.strftime('%d/%m/%Y')}"""
        
    except Exception as e:
        return """😅 **Ops! Tive um pequeno problema, mas não desista de mim! 🤗**

**🔄 Vamos tentar de novo?** Me faça uma pergunta simples sobre:
• Salas disponíveis
• Localização de laboratórios  
• Software e equipamentos
• Contato do SENAI

**💬 Exemplo:** *"Preciso de uma sala com computadores"*

Estou aqui para te ajudar! ✨"""

# ========= NEW INTELLIGENT FUNCTIONS - NO MORE GENERIC ERRORS =========

def get_rooms_location_info_smart(classrooms):
    """Return location information about classrooms with smart handling"""
    from models import Classroom
    from app import db
    
    try:
        all_classrooms = db.session.query(Classroom).all()
        if not all_classrooms:
            return "🏢 **SENAI Morvan Figueiredo**\nEntre em contato com a secretaria para informações sobre localização das salas! 📞"
        
        response = "🗺️ **Localização das Salas - SENAI Morvan Figueiredo:**\n\n"
        
        # Group by blocks
        blocks = {}
        for room in all_classrooms:
            block = room.block or 'Sem bloco definido'
            if block not in blocks:
                blocks[block] = []
            blocks[block].append(room)
        
        for block_name, rooms in sorted(blocks.items()):
            response += f"🏢 **{block_name}:**\n"
            for room in sorted(rooms, key=lambda x: x.name):
                response += f"  • **{room.name}** - {room.capacity} pessoas"
                if room.has_computers:
                    response += " 💻"
                response += "\n"
            response += "\n"
        
        response += "📍 **Endereço:** SENAI Morvan Figueiredo\n"
        response += "💬 **Contato:** Solicite no sistema para orientações detalhadas"
        response += get_question_menu()
        
        return response
        
    except Exception as e:
        import logging
        logging.error(f"Error in get_rooms_location_info_smart: {str(e)}")
        return "🗺️ **Localização:** As salas estão distribuídas em diferentes blocos do SENAI Morvan Figueiredo. Entre em contato com a secretaria para localização exata! 📞"

def get_schedule_info_smart(classrooms, schedules):
    """Return schedule information with smart handling"""
    from datetime import datetime
    import pytz
    
    try:
        sp_tz = pytz.timezone('America/Sao_Paulo')
        current_time = datetime.now(sp_tz)
        
        response = f"📅 **Informações de Horários - {current_time.strftime('%d/%m/%Y')}:**\n\n"
        
        if schedules:
            # Group by shifts
            shifts = {'morning': [], 'afternoon': [], 'night': [], 'fullday': []}
            shift_names = {
                'morning': 'Manhã (8h-12h)',
                'afternoon': 'Tarde (13h-17h)', 
                'night': 'Noite (18h-22h)',
                'fullday': 'Período Integral'
            }
            
            for schedule in schedules:
                if hasattr(schedule, 'shift') and schedule.shift in shifts:
                    shifts[schedule.shift].append(schedule)
            
            for shift_key, shift_schedules in shifts.items():
                if shift_schedules:
                    response += f"🕓 **{shift_names[shift_key]}:**\n"
                    for schedule in shift_schedules[:5]:  # Limit to 5 per shift
                        classroom_name = "Sala desconhecida"
                        if hasattr(schedule, 'classroom_id'):
                            classroom = next((c for c in classrooms if c.id == schedule.classroom_id), None)
                            if classroom:
                                classroom_name = classroom.name
                        
                        response += f"  • **{classroom_name}** - {schedule.course_name}\n"
                    response += "\n"
        else:
            response += "📋 **Nenhum agendamento encontrado para hoje.**\n\n"
        
        response += "🕰️ **Horários de Funcionamento:**\n"
        response += "  • Segunda a Sexta: 7h30 às 22h\n"
        response += "  • Sábado: 7h30 às 12h\n"
        response += "\n💬 **Para agendamentos:** Solicite no sistema"
        response += get_question_menu()
        
        return response
        
    except Exception as e:
        import logging
        logging.error(f"Error in get_schedule_info_smart: {str(e)}")
        return "📅 **Horários de Funcionamento:**\nSegunda a Sexta: 7h30-22h | Sábado: 7h30-12h\n📞 Entre em contato com a secretaria para agendamentos!"

def get_analytics_and_trends_smart(classrooms, schedules, current_time):
    """Return analytics with smart handling"""
    from datetime import datetime
    import pytz
    
    try:
        sp_tz = pytz.timezone('America/Sao_Paulo')
        current_sp_time = datetime.now(sp_tz)
        
        total_rooms = len(classrooms) if classrooms else 0
        total_capacity = sum(room.capacity for room in classrooms if hasattr(room, 'capacity')) if classrooms else 0
        rooms_with_computers = len([r for r in classrooms if hasattr(r, 'has_computers') and r.has_computers]) if classrooms else 0
        
        response = f"📊 **Análise do Sistema - {current_sp_time.strftime('%d/%m/%Y às %H:%M')}:**\n\n"
        response += f"🏫 **Estrutura Geral:**\n"
        response += f"  • Total de salas: **{total_rooms}**\n"
        response += f"  • Capacidade total: **{total_capacity} pessoas**\n"
        response += f"  • Salas com computadores: **{rooms_with_computers}**\n\n"
        
        if classrooms:
            # Capacity distribution
            small = len([r for r in classrooms if hasattr(r, 'capacity') and r.capacity <= 20])
            medium = len([r for r in classrooms if hasattr(r, 'capacity') and 21 <= r.capacity <= 35])
            large = len([r for r in classrooms if hasattr(r, 'capacity') and r.capacity > 35])
            
            response += "📈 **Distribuição por Tamanho:**\n"
            response += f"  • Pequenas (até 20): **{small} salas**\n"
            response += f"  • Médias (21-35): **{medium} salas**\n"
            response += f"  • Grandes (36+): **{large} salas**\n\n"
        
        active_schedules = len(schedules) if schedules else 0
        response += f"🗂️ **Agendamentos:** {active_schedules} atividades programadas\n\n"
        
        response += "🎯 **Recomendações:**\n"
        response += "  • Consulte disponibilidade em tempo real\n"
        response += "  • Reserve com antecedência para garantir vaga\n"
        response += "  • Considere horários alternativos se necessário"
        response += get_question_menu()
        
        return response
        
    except Exception as e:
        import logging
        logging.error(f"Error in get_analytics_and_trends_smart: {str(e)}")
        return f"📊 **Sistema SENAI Morvan Figueiredo:**\nTotal de salas disponíveis para consulta\n📞 Entre em contato para mais detalhes sobre ocupação e agendamentos!"

def get_basic_classroom_info(classrooms):
    """Return basic classroom information as fallback"""
    try:
        if not classrooms:
            return "🏫 **Sistema SENAI Morvan Figueiredo**\nEntre em contato com a secretaria para informações sobre as salas! 📞"
        
        total_rooms = len(classrooms)
        total_capacity = sum(room.capacity for room in classrooms if hasattr(room, 'capacity'))
        
        response = f"🏫 **Salas Disponíveis ({total_rooms} salas):**\n\n"
        
        for room in classrooms[:4]:  # Show first 4 rooms
            response += f"• **{room.name}** ({getattr(room, 'block', 'Bloco n/d')})\n"
            response += f"  👥 {getattr(room, 'capacity', 0)} pessoas"
            if hasattr(room, 'has_computers') and room.has_computers:
                response += " | 💻 Computadores"
            response += "\n\n"
        
        if len(classrooms) > 4:
            response += f"... e mais {len(classrooms) - 4} salas\n\n"
        
        response += f"📊 **Total:** {total_capacity} pessoas\n"
        response += "💬 **Mais informações:** Solicite no sistema"
        response += get_question_menu()
        
        return response
        
    except Exception:
        return "🏫 **SENAI Morvan Figueiredo**\nSistema de salas disponível. Entre em contato com a secretaria! 📞"

def get_all_software_options(classrooms):
    """Return all available software options as fallback"""
    try:
        if not classrooms:
            return "💻 **Software Disponível**\nEntre em contato com a secretaria para informações sobre software! 📞"
        
        software_rooms = [room for room in classrooms if hasattr(room, 'software') and room.software]
        
        if not software_rooms:
            return "💻 **Software nas Salas:**\nInformações sendo atualizadas. Consulte a secretaria! 📞"
        
        response = "💻 **Software Disponível nas Salas:**\n\n"
        
        for room in software_rooms:
            response += f"• **{room.name}** ({getattr(room, 'block', 'Bloco n/d')})\n"
            response += f"  🛠️ {room.software}\n\n"
        
        response += "💬 **Para usar:** Solicite no sistema"
        response += get_question_menu()
        
        return response
        
    except Exception:
        return "💻 **Software:**\nUnity, Blender, Visual Studio, Office e mais\n📞 Consulte disponibilidade na secretaria!"

def get_basic_capacity_info(classrooms):
    """Return basic capacity information as fallback"""
    try:
        if not classrooms:
            return "👥 **Capacidade das Salas**\nSolicite no sistema! 💬"
        
        response = "👥 **Capacidade das Salas:**\n\n"
        
        # Show simplified info
        for room in classrooms[:4]:
            if hasattr(room, 'capacity'):
                response += f"• **{room.name}**: {room.capacity} pessoas\n"
        
        total_capacity = sum(room.capacity for room in classrooms if hasattr(room, 'capacity'))
        response += f"\n📊 **Total:** {total_capacity} pessoas em {len(classrooms)} salas"
        
        return response
        
    except Exception:
        return "👥 **Salas:** 20-40 pessoas por sala\n💬 Solicite no sistema!"

def get_question_menu():
    """Generate a short menu of questions the user can ask"""
    return "\n\n❓ **Mais ajuda?** Pergunte sobre salas, software ou horários!"

def get_emergency_helpful_response(user_message, classrooms):
    """Emergency fallback that always provides something useful"""
    from datetime import datetime
    import pytz
    
    sp_tz = pytz.timezone('America/Sao_Paulo')
    current_time = datetime.now(sp_tz)
    
    total_rooms = len(classrooms) if classrooms else 0
    
    return f"""🤖 **Olá! Sou o assistente do SENAI Morvan Figueiredo! 😊**

Percebi que você disse: *"{user_message}"*

**🏫 Informações Rápidas:**
• {total_rooms} salas disponíveis
• Horário atual: {current_time.strftime('%H:%M')}
• Funcionamento: Segunda a Sexta (7h30-22h)

**💬 Posso te ajudar com:**
• "Que salas estão livres agora?"
• "Preciso de uma sala para X pessoas"
• "Onde fica a [nome da sala]?"
• "Que software tem disponível?"

**💬 Solicite no sistema** - Para reservas e agendamentos

**🤝 Como posso te ajudar de verdade?** Me faça uma pergunta mais específica! ✨{get_question_menu()}"""